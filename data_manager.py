"""
data_manager.py
Centralized module to parse Excel master lists during application boot.
Provides O(1) dictionary lookups for Excel mapping and formatted string injection for AI prompts.
"""
import os
import re
import logging
from path_utils import get_asset_path

# Paths resolving via utility (Handles development vs EXE root)
_DATA_DIR        = get_asset_path("Data structure")
_SUPPLIER_FILE   = os.path.join(_DATA_DIR, "Danh sach nha cung cap.xlsx")
_INGREDIENT_FILE = os.path.join(_DATA_DIR, "Danh mục vật tư, hàng hoá.xlsx")
_KHO_FILE        = os.path.join(_DATA_DIR, "Danh sach kho.xlsx")
_DEFAULT_ALIAS_FILE = os.path.join(_DATA_DIR, "Tu_dien_alias.csv")  # module-level fallback only

# Schema constants for new ingredient file (Danh mục vật tư, hàng hoá.xlsx)
# Row 1-4: Công ty, tiêu đề | Row 5: Column headers | Row 6+: Data
_INGREDIENT_DATA_ROW = 6
# Column indices (0-based) in the new file:
_COL_MA_VT    = 1   # Ma_Vt:    Mã vật tư
_COL_TEN_VT   = 2   # Ten_Vt:   Tên vật tư (có tiền tố nhóm, vd: BANH-Baking powder)
_COL_DVT      = 4   # Dvt:      Đơn vị tính chính
_COL_GIA_MUA  = 6   # Gia_Mua:  Giá mua tham chiếu theo Dvt chính
_COL_DVT0     = 9   # Dvt0:     Đơn vị quy đổi
_COL_HE_SO0   = 10  # He_So0:   Hệ số quy đổi (1 Dvt = He_So0 Dvt0)
_COL_GIA_MUA0 = 12  # Gia_Mua0: Giá mua tham chiếu theo Dvt0
_COL_MA_NH_VT = 32  # Ma_Nh_Vt: Mã nhóm vật tư (dùng để xác định bộ phận BEP/BAR/BANH/RANG)

# Schema: minimum number of data rows expected to consider a file "healthy"
_MIN_SUPPLIER_ROWS   = 5
_MIN_INGREDIENT_ROWS = 10


def validate_data_files() -> list[str]:
    """
    Kiểm tra tính toàn vẹn của các file Data structure khi app khởi động.
    Trả về danh sách chuỗi cảnh báo (warnings). Nếu rỗng → mọi thứ ổn.
    
    Được thiết kế để gọi 1 lần từ main_app_qt.py ngay sau khi QApplication khởi tạo.
    """
    import openpyxl
    warnings = []

    # --- Kiểm tra thư mục Data structure ---
    if not os.path.isdir(_DATA_DIR):
        warnings.append(f"❌ THIẾU THƯ MỤC: Không tìm thấy '{_DATA_DIR}'. "
                        f"Tạo thư mục và đặt file Excel vào trong.")
        return warnings  # Không cần kiểm tra tiếp

    # --- Kiểm tra file Nhà Cung Cấp ---
    if not os.path.isfile(_SUPPLIER_FILE):
        warnings.append(f"❌ THIẾU FILE NCC: '{os.path.basename(_SUPPLIER_FILE)}' "
                        f"không tồn tại trong thư mục Data structure.")
    else:
        try:
            wb = openpyxl.load_workbook(_SUPPLIER_FILE, data_only=True, read_only=True)
            ws = wb.active
            count = sum(1 for row in ws.iter_rows(min_row=2, values_only=True)
                        if row[1] and row[2])
            wb.close()
            if count < _MIN_SUPPLIER_ROWS:
                warnings.append(f"⚠️ DỮ LIỆU THƯA: '{os.path.basename(_SUPPLIER_FILE)}' "
                                 f"chỉ có {count} nhà cung cấp (tối thiểu kỳ vọng: {_MIN_SUPPLIER_ROWS}).")
        except Exception as e:
            warnings.append(f"❌ LỖI ĐỌC FILE NCC: {os.path.basename(_SUPPLIER_FILE)} — {e}")

    # --- Kiểm tra file Danh mục Vật tư ---
    if not os.path.isfile(_INGREDIENT_FILE):
        warnings.append(f"❌ THIẾU FILE VẬT TƯ: '{os.path.basename(_INGREDIENT_FILE)}' "
                        f"không tồn tại trong thư mục Data structure.")
    else:
        try:
            wb = openpyxl.load_workbook(_INGREDIENT_FILE, data_only=True, read_only=True)
            ws = wb.active
            # Data bắt đầu từ row 6 (rows 1-4 tiêu đề công ty, row 5 là header cột)
            count = sum(1 for row in ws.iter_rows(min_row=_INGREDIENT_DATA_ROW, values_only=True)
                        if row[_COL_MA_VT] and row[_COL_TEN_VT])
            wb.close()
            if count < _MIN_INGREDIENT_ROWS:
                warnings.append(f"⚠️ DỮ LIỆU THƯA: '{os.path.basename(_INGREDIENT_FILE)}' "
                                 f"chỉ có {count} vật tư (tối thiểu kỳ vọng: {_MIN_INGREDIENT_ROWS}).")
        except Exception as e:
            warnings.append(f"❌ LỖI ĐỌC FILE VẬT TƯ: {os.path.basename(_INGREDIENT_FILE)} — {e}")

    return warnings


import json

# Model Aliases & Defaults (April 2026)
# NOTE: These are in-code fallback defaults only.
# Active values are always loaded from lighthouse_config.json at runtime.
# Keep in sync with the defaults written to lighthouse_config.json.
_DEFAULT_MODELS = {
    "light_primary":  "gemini-3.1-flash-lite-preview",
    "light_fallback": "gemini-2.5-flash",
    "pro_primary":    "gemini-2.5-pro",
    "pro_fallback":   "gemini-2.5-pro"
}

class DataManager:
    def __init__(self):
        self.suppliers_dict = {}
        self.suppliers_context_str = ""

        self.items_dict = {}
        self.items_by_name = {}
        self.items_context_str = ""
        self.kho_dict = {}  # {dept_upper -> ma_kho}, e.g. {"BAR": "LH-BAR", "BEP": "LH-BEP", ...}
        self.is_loaded = False
        self._data_version = 0
        
        # Model Configs
        from path_utils import get_root_dir
        self.models = _DEFAULT_MODELS.copy()
        self.config_file = os.path.join(get_root_dir(), "lighthouse_config.json")

    def load_config(self):
        """Tải cấu hình model từ file bên ngoài (không nén trong exe)."""
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, "r", encoding="utf-8") as f:
                    cfg = json.load(f)
                    if "models" in cfg:
                        self.models.update(cfg["models"])
            except Exception as e:
                print(f"Warning: Could not load model config: {e}")

    def save_config(self, api_key=None, admin_pass=None, models=None):
        """Lưu cấu hình (HWID, API Key, Password đã được mã hóa ở main_app_qt) + Models."""
        # Chú ý: Việc mã hóa api_key/pass vẫn thực hiện ở main_app_qt để giữ logic security tập trung.
        # Hàm này chủ yếu để cập nhật phần 'models' trong file json.
        current_data = {}
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, "r", encoding="utf-8") as f:
                    current_data = json.load(f)
            except: pass
            
        if models:
            current_data["models"] = models
            self.models.update(models)
        
        # Các trường khác được pass trực tiếp từ AdminConfigDialog (dạng đã obscure)
        if api_key: current_data["api_key"] = api_key
        if admin_pass: current_data["admin_password"] = admin_pass
        
        try:
            with open(self.config_file, "w", encoding="utf-8") as f:
                json.dump(current_data, f, indent=2)
        except Exception as e:
            print(f"Error saving config: {e}")

    def get_alias_file_path(self) -> str:
        """Returns the alias CSV path from lighthouse_config.json, falling back to the default asset path."""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    cfg = json.load(f)
                custom = cfg.get('alias_file_path', '')
                if custom and os.path.isfile(custom):
                    return custom
        except Exception:
            pass
        return get_asset_path(os.path.join('Data structure', 'Tu_dien_alias.csv'))

    def save_alias_file_path(self, path: str):
        """Persists a custom alias file path to lighthouse_config.json."""
        current_data = {}
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    current_data = json.load(f)
            except Exception:
                pass
        current_data['alias_file_path'] = path
        try:
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(current_data, f, indent=2, ensure_ascii=False)
        except Exception as e:
            logging.error(f"[DataManager] Lỗi lưu alias_file_path: {e}")

    def list_available_models(self, api_key: str) -> list[str]:
        """Gọi API để lấy danh sách các model Gemini khả dụng."""
        if not api_key: return []
        try:
            from google import genai
            client = genai.Client(api_key=api_key)
            model_list = []
            for m in client.models.list():
                # Lọc lấy các model gemini
                if "gemini" in m.name.lower():
                    # Format thường là "models/gemini-..." -> lấy phần sau models/
                    name = m.name.split("/")[-1]
                    model_list.append(name)
            return sorted(list(set(model_list)))
        except Exception as e:
            print(f"Error listing models: {e}")
            return []

    def should_use_minimal_thinking(self, model_name: str, is_primary: bool = True) -> bool:
        """
        Kiểm tra xem model có nên dùng Minimal Thinking (thinking_budget=0) không.
        
        Quy tắc:
        - Chỉ áp dụng cho PRIMARY slot (is_primary=True).
        - Fallback models (is_primary=False) KHÔNG được gửi budget=0.
        - Model dòng PRO KHÔNG được gửi budget=0 vì chúng bắt buộc phải có thinking budget.
        - Chỉ áp dụng cho các đời model Flash/Lite mới (2.0, 2.5, 3.x).
        """
        if not model_name or not is_primary: return False
        name = model_name.lower()
        
        # Dòng Pro không hỗ trợ budget=0
        if "pro" in name: return False
        
        # Các đời mới hỗ trợ/cần set thinking config để tối ưu
        new_generations = ["2.0", "2.5", "3.0", "3.1", "3.x", "3-"]
        return any(gen in name for gen in new_generations)

    def load_all(self):
        """Loads both Excel files into memory. Should be called exactly once at startup."""
        try:
            self._parse_suppliers()
            self._parse_ingredients()
            self._parse_kho()
            self._parse_aliases()
            self._data_version += 1
            self.is_loaded = True
        except Exception as e:
            raise RuntimeError(f"Lỗi khi đọc file Master List: {e}")

    def _parse_kho(self):
        """Parse Danh sach kho.xlsx → {dept_upper: ma_kho}.
        Expected columns: STT | Tên bộ phận | Mã Kho
        """
        import openpyxl
        self.kho_dict = {}
        if not os.path.isfile(_KHO_FILE):
            logging.warning(f"[DataManager] Không tìm thấy file kho: {_KHO_FILE}")
            return
        try:
            wb = openpyxl.load_workbook(_KHO_FILE, data_only=True, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                dept = row[1]  # Cột B: Tên bộ phận
                ma_kho = row[2]  # Cột C: Mã Kho
                if dept and ma_kho:
                    self.kho_dict[str(dept).strip().upper()] = str(ma_kho).strip()
            wb.close()
            logging.info(f"[DataManager] Đã nạp {len(self.kho_dict)} mục mã kho: {self.kho_dict}")
        except Exception as e:
            logging.warning(f"[DataManager] Lỗi đọc file kho: {e}")

    def _parse_aliases(self):
        """Parse Tu_dien_alias.csv -> {alias_lower: {'code': code, 'units': [{'unit': unit, 'factor': factor}, ...]}}
        Logic: Multiple units for the same alias are supported. 
        Records with same (alias, slang_unit) are deduplicated (latest wins).
        """
        self.aliases_dict = {}
        alias_file = self.get_alias_file_path()
        if not os.path.isfile(alias_file):
            return
        import csv
        try:
            # Step 1: Gather all records, latest record for (alias, unit) pair wins
            temp_dedup = {} # key: (alias_lower, unit_lower)

            with open(alias_file, mode='r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                next(reader, None) # Skip header
                for row in reader:
                    # Enforce 6-column schema: 0:Alias, 1:Ma, 2:Ten, 3:Dvt, 4:Dvt_long, 5:He_so
                    if len(row) >= 6:
                        alias = str(row[0]).strip().lower()
                        code = str(row[1]).strip()
                        alias_unit = str(row[4]).strip()
                        raw_factor = str(row[5]).strip()
                        
                        if not alias or not code:
                            continue

                        alias_factor = None
                        if raw_factor:
                            try:
                                alias_factor = float(raw_factor)
                            except ValueError:
                                pass
                        
                        # Use (alias, unit) as dedup key
                        unit_key = alias_unit.lower()
                        temp_dedup[(alias, unit_key)] = {
                            "code": code,
                            "unit": alias_unit,
                            "factor": alias_factor
                        }

            # Step 2: Group by alias
            for (alias, _), info in temp_dedup.items():
                if alias not in self.aliases_dict:
                    self.aliases_dict[alias] = {"code": info["code"], "units": []}
                
                self.aliases_dict[alias]["units"].append({
                    "unit": info["unit"],
                    "factor": info["factor"]
                })
                
            logging.info(f"[DataManager] Đã nạp {len(self.aliases_dict)} unique aliases (với đa đơn vị).")
        except Exception as e:
            logging.warning(f"[DataManager] Lỗi đọc file alias: {e}")

    def save_alias(self, alias: str, code: str, ref_name: str = "", master_unit: str = "", alias_unit: str = "", alias_factor: float = None):
        """Lưu một alias mới vào file CSV và cập nhật RAM."""
        alias_key = alias.strip().lower()
        if not alias_key or not code: return
        
        # Cập nhật RAM theo cấu trúc 1-N
        if alias_key not in self.aliases_dict:
            self.aliases_dict[alias_key] = {"code": code.strip(), "units": []}
        
        # Check if this unit already exists in RAM to update it, else append
        unit_found = False
        target_unit_clean = alias_unit.strip().lower()
        for u_info in self.aliases_dict[alias_key]["units"]:
            if u_info["unit"].strip().lower() == target_unit_clean:
                u_info["factor"] = alias_factor
                unit_found = True
                break
        
        if not unit_found:
            self.aliases_dict[alias_key]["units"].append({
                "unit": alias_unit.strip(),
                "factor": alias_factor
            })
        
        # Ghi vào file CSV
        import csv
        alias_file = self.get_alias_file_path()
        file_exists = os.path.isfile(alias_file)
        try:
            # Nếu file chưa có, tạo mới với header 6 cột
            if not file_exists:
                os.makedirs(os.path.dirname(alias_file), exist_ok=True)
                with open(alias_file, mode='w', encoding='utf-8', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(['Alias (Tên trên phiếu)', 'Mã vật tư', 'Tên vật tư', 'Đơn vị tính', 'Đơn vị tính lóng', 'Hệ số lóng'])

            with open(alias_file, mode='a', encoding='utf-8', newline='') as f:
                writer = csv.writer(f)
                factor_str = str(alias_factor) if alias_factor is not None else ""
                writer.writerow([alias.strip(), code.strip(), ref_name.strip(), master_unit.strip(), alias_unit.strip(), factor_str])
            logging.info(f"[DataManager] Đã lưu alias: '{alias}' -> Mã: '{code}' (Master Unit: '{master_unit}', Slang Unit: '{alias_unit}', Factor: {alias_factor})")
        except Exception as e:
            logging.error(f"[DataManager] Lỗi ghi file alias: {e}")

    def _parse_suppliers(self):
        import openpyxl
        try:
            wb = openpyxl.load_workbook(_SUPPLIER_FILE, data_only=True)
            ws = wb.active
            parts = []
            dict_mapping = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                code = row[1]
                name = row[2]
                products = row[3]  # Cột D: Sản phẩm mẫu
                if code and name:
                    code_str = str(code).strip()
                    name_str = str(name).strip()
                    prod_str = f" (Sản phẩm: {str(products).strip()})" if products else ""
                    parts.append(f"{code_str}={name_str}{prod_str}")
                    dict_mapping[code_str] = name_str
            self.suppliers_context_str = " | ".join(parts)
            self.suppliers_dict = dict_mapping
        except Exception as e:
            self.suppliers_context_str = f"[SUPPLIER LIST UNAVAILABLE: {e}]"
            self.suppliers_dict = {}

    def _parse_ingredients(self):
        """
        Đọc file 'Danh mục vật tư, hàng hoá.xlsx' (phiên bản mới từ EZSoft).
        Cấu trúc file:
          Row 1-4: Tiêu đề công ty
          Row 5:   Header cột (Ma_Vt, Ten_Vt, Dvt, Dvt0, He_So0, Ma_Nh_Vt, ...)
          Row 6+:  Dữ liệu
        Cột sử dụng (0-based index):
          [1]  Ma_Vt    — Mã vật tư
          [2]  Ten_Vt   — Tên (có tiền tố nhóm, vd: BANH-Baking powder)
          [4]  Dvt      — Đơn vị tính chính
          [9]  Dvt0     — Đơn vị quy đổi
          [10] He_So0   — Hệ số quy đổi (1 Dvt = He_So0 Dvt0)
          [32] Ma_Nh_Vt — Mã nhóm vật tư (BEP/BAR/BANH/RANG/CCDC...)
        """
        import openpyxl
        try:
            wb = openpyxl.load_workbook(_INGREDIENT_FILE, data_only=True)
            ws = wb.active
            ref = {}
            by_name = {}
            by_code = {}
            ocr_parts = []

            for row in ws.iter_rows(min_row=_INGREDIENT_DATA_ROW, values_only=True):
                code    = row[_COL_MA_VT]
                name    = row[_COL_TEN_VT]
                unit    = row[_COL_DVT]
                gia_mua = row[_COL_GIA_MUA]   # Giá mua theo Dvt chính
                dvt0    = row[_COL_DVT0]
                he_so   = row[_COL_HE_SO0]
                gia_mua0= row[_COL_GIA_MUA0]  # Giá mua theo Dvt0
                group   = row[_COL_MA_NH_VT]

                if not name:
                    continue

                # Bóc tiền tố nhóm khỏi tên: "BANH-Baking powder" → "Baking powder"
                # Hỗ trợ cả dạng có số: "BEP01-Tương cà Heinz" → "Tương cà Heinz"
                # Hỗ trợ CCDC: "CCDCBAR-Ly thuỷ tinh" → "Ly thuỷ tinh"
                display = str(name).strip()
                stripped = re.sub(r'^(CCDC[A-Z]*|BEP|BAR|BANH|RANG)\d*[-_]', '', display, flags=re.IGNORECASE)
                if stripped != display:
                    logging.debug(f"[DataManager] Strip prefix: '{display}' → '{stripped}'")
                    display = stripped

                # Context string cho LLM prompt: Tên/DVT
                ocr_parts.append(f"{display}/{unit}" if unit else display)

                # Key là display.lower() để tra cứu O(1) sau khi mapping
                # Xử lý giá mua an toàn (Gia_Mua đi với Dvt, Gia_Mua0 đi với Dvt0)
                def _safe_float(val):
                    if val is None or str(val).strip().lower() in ("", "none"):
                        return None
                    try: return float(val)
                    except: return None

                item_record = {
                    "code":     str(code).strip() if code else "",
                    "name":     display,
                    "unit":     str(unit).strip() if unit else "",
                    "gia_mua":  _safe_float(gia_mua),
                    "dvt0":     str(dvt0).strip() if dvt0 else "",
                    "he_so0":   he_so if he_so else None,
                    "gia_mua0": _safe_float(gia_mua0),
                    "group":    str(group).strip() if group else "",
                }
                name_key = display.lower()
                by_name.setdefault(name_key, []).append(item_record)
                
                # Cập nhật items_dict (vẫn giữ map 1-1 làm fallback, cái sau đè cái trước nếu trùng tên)
                ref[name_key] = item_record

                # Cập nhật tra cứu theo Mã vật tư
                if code:
                    by_code[str(code).strip().lower()] = item_record

            self.items_context_str = " | ".join(ocr_parts)
            self.items_dict = ref
            self.items_by_name = by_name
            self.items_by_code = by_code
        except Exception as e:
            self.items_context_str = f"[INGREDIENT LIST UNAVAILABLE: {e}]"
            self.items_dict = {}
            self.items_by_name = {}
            self.items_by_code = {}


# Global singleton
app_data = DataManager()
