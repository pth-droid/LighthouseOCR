"""
core_excel_mapper.py
Maps a list of canonical invoice JSON dicts into a fresh output Excel file.

Appends data dynamically into a copy of `PNMH-Mẫu phiếu nhập mua hàng.xlsx`
"""

import os
import difflib
import shutil
import json
import logging
from datetime import datetime
from typing import List, Dict, Any, Tuple

import google.genai as genai
from google.genai import types

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation

from voucher_manager import get_next_voucher_number
from core_rate_limiter import global_rate_limiter
from constants import HIGH_RISK_CONF_THRESHOLD
from path_utils import get_root_dir

# ---- Paths ----
_ROOT_DIR = get_root_dir()
_TEMPLATE_FILE = os.path.join(_ROOT_DIR, "Data structure", "PNMH-Mẫu phiếu nhập mua hàng.xlsx")
_CHIPHI_TEMPLATE_FILE = os.path.join(_ROOT_DIR, "Data structure", "Mẫu Nhập Chi phí.xlsx")

# ---- Color palette ----
_FILL_NONE      = PatternFill(fill_type=None)
_FILL_PALE_BLUE = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
_FILL_RED       = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
_FILL_YELLOW    = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_FILL_HEADER    = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")

_FONT_HEADER    = Font(color="FFFFFF", bold=True, size=10)
_FONT_DATA      = Font(size=14)
_THIN_BORDER    = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

_TOTAL_COLUMNS = 27
_DATA_START_ROW = 3
_NOTES_CELL_COMMENT = (
    "QUY TẮC MÀU SẮC & TÍNH TOÁN:\n"
    "----------------------------------\n"
    "🔴 ĐỎ: Hóa đơn rủi ro cao (Confidence thấp hoặc thiếu NCC).\n"
    "🟡 VÀNG: Vật tư chưa khớp Master List hoặc Sai số liệu (Qty * Price != Total).\n"
    "🔵 XANH NHẠT: Hóa đơn bình thường (tô xen kẽ để dễ nhìn).\n\n"
    "LOGIC TÍNH GIÁ:\n"
    "- Đơn giá (cột Q,R): Đã bao gồm VAT và Chiết khấu/Phí vận chuyển phân bổ.\n"
    "- Phân bổ: Chiết khấu/Phí cấp đơn được chia theo tỷ trọng giá trị từng dòng hàng.\n"
    "- VAT: Được tính trên giá sau chiết khấu.\n"
    "- Giá gốc: Xem chi tiết phép tính điều chỉnh trong nội dung ô ghi chú."
)


# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------

def _get_fill_for_invoice(invoice_idx: int, is_high_risk: bool) -> PatternFill:
    if is_high_risk:
        return _FILL_RED
    return _FILL_NONE if invoice_idx % 2 == 0 else _FILL_PALE_BLUE


# ---- Đᨋn vị tính aliases (dùng cho quy đổi DVT0 → DVT) ----
# Group các cách viết khác nhau cùng chỉ một đơn vị
_UNIT_ALIAS_GROUPS = [
    {"kg", "kilogram", "kilo", "kgs"},
    {"gram", "gr", "g", "gam"},
    {"lít", "lit", "liter", "litre", "l"},
    {"ml", "mililit", "cc"},
    {"hộp", "hop", "box", "h.p"},
    {"lon", "can", "l.n"},
    {"chai", "bottle", "chai/c"},
    {"gói", "goi", "pack", "túi"},
    {"thùng", "thung", "carton", "ctn"},
    {"cái", "cai", "piece", "pcs"},
    {"cuộn", "cuon", "roll"},
]

def _match_unit(invoice_unit: str, master_dvt0: str) -> bool:
    """
    Kiểm tra xem đơn vị trên hoá đơn có tương đương Dvt0 trong Master List không.
    Hỗ trợ các cách viết khác nhau (kg/Kg/KG, hộp/Hop/box, ...).
    """
    if not invoice_unit or not master_dvt0:
        return False
    u1 = invoice_unit.strip().lower()
    u2 = master_dvt0.strip().lower()
    if u1 == u2:
        return True
    for group in _UNIT_ALIAS_GROUPS:
        if u1 in group and u2 in group:
            return True
    return False


def _extract_size_from_name(name: str):
    """
    Trích xuất kích thước/trọng lượng/thể tích ghi trong tên mặt hàng.
    Trả về (value: float, unit_str: str) hoặc (None, "") nếu không tìm thấy.

    Hỗ trợ:
      - Có hoặc không có khoảng trắng: "950ml", "950 ml", "1.5 L", "1,5L"
      - Số thập phân với dấu phẩy hoặc chấm: "1,5L" → 1.5L
      - Tiếng Việt: "lít", "lít", "gam", "ki-lô-gam"
      - Đơn vị thể tích: ml, cc, L, lít, litre, liter
      - Đơn vị khối lượng: g, gr, gam, gram, mg, kg, kgs, oz, lb, pound
    Ví dụ:
      "Sữa tươi Meiji 950ml"    → (950.0, "ml")
      "Sữa tươi Meiji 950 ml"   → (950.0, "ml")
      "Phô mai dairymont 2kg"   → (2.0,   "kg")
      "Nước Lavie 1,5L"         → (1.5,   "L")
      "Nước Lavie 1.5 lít"      → (1.5,   "L")
      "Kem sữa 500cc"           → (500.0, "ml")  # cc → ml
      "Bơ Anchor 500g"          → (500.0, "g")
      "Bột cacao 200 gram"      → (200.0, "g")
    """
    # Mỗi pattern: (regex, canonical_unit)
    # Thứ tự quan trọng: đơn vị dài hơn phải đứng trước để tránh partial match
    # VD: "kg" phải trước "g" để "2kg" không bị đọc nhầm là "2k" + "g"
    patterns = [
        # ── Thể tích ──
        (r'(\d+(?:[.,]\d+)?)\s*cc\b',                          'ml'),   # cc ≡ ml
        (r'(\d+(?:[.,]\d+)?)\s*ml\b',                          'ml'),
        (r'(\d+(?:[.,]\d+)?)\s*(?:lít|lít|litre|liter|lit)\b', 'L'),   # tiếng Việt có dấu + không dấu
        (r'(\d+(?:[.,]\d+)?)\s*L\b',                           'L'),   # viết hoa (1L, 2L)
        # ── Khối lượng ──
        (r'(\d+(?:[.,]\d+)?)\s*(?:kg|kgs?|ki.?lô.?gam)\b',   'kg'),  # kg, kgs, ki-lô-gam
        (r'(\d+(?:[.,]\d+)?)\s*(?:mg)\b',                      'mg'),
        (r'(\d+(?:[.,]\d+)?)\s*(?:gram|gam|gr)\b',             'g'),   # dài hơn trước "g"
        (r'(\d+(?:[.,]\d+)?)\s*g\b',                           'g'),   # đứng sau để tránh "mg" bị bắt
        (r'(\d+(?:[.,]\d+)?)\s*(?:pound|lb)\b',                'lb'),
        (r'(\d+(?:[.,]\d+)?)\s*oz\b',                          'oz'),
    ]
    name_str = str(name or "")
    for pat, unit_key in patterns:
        m = re.search(pat, name_str, re.IGNORECASE)
        if m:
            try:
                val = float(m.group(1).replace(',', '.'))
                if val > 0:
                    return val, unit_key
            except ValueError:
                continue
    return None, ""


# Map đơn vị volume/weight về nhóm để so sánh dimension
_VOLUME_UNITS = {"ml", "l", "lít", "lit", "liter", "litre", "cc"}
_WEIGHT_UNITS = {"g", "gr", "gam", "gram", "mg", "kg", "kgs", "oz", "lb", "pound"}

# Từ khoá nhận dạng sản phẩm sữa nước
# Lighthouse áp dụng quy ước 1ml = 1g cho nhóm này (mật độ sữa ≈ 1.03 g/ml, làm tròn = 1)
_DAIRY_LIQUID_KEYWORDS = [
    "sữa", "sua", "milk", "sữa tươi", "sữa hộp", "sữa lon",
    "sữa đặc", "sữa không đường", "sữa ít béo", "sữa nguyên kem",
    "fresh milk", "whole milk", "skim milk",
    # Phổ biến tại Lighthouse:
    "meiji", "vinamilk", "dalat milk", "meadow fresh", "greenfields",
    "anchor milk", "dutch lady", "friso", "similac",
]


def _is_dairy_liquid(name: str) -> bool:
    """
    Nhận dạng sản phẩm sữa nước.
    Lighthouse áp dụng quy ước 1ml ≈ 1g cho nhóm này.
    """
    name_lower = name.lower()
    return any(kw in name_lower for kw in _DAIRY_LIQUID_KEYWORDS)


def _convert_volume_to_weight_dairy(size_val: float, size_unit: str, target_dvt: str) -> float | None:
    """
    Quy đổi thể tích → khối lượng cho sản phẩm sữa nước.
    Quy ước Lighthouse: 1ml = 1g, 1L = 1kg.

    Args:
        size_val:   Giá trị kích thước (VD: 950, 1.5)
        size_unit:  Đơn vị trong tên hàng ("ml", "L", "cc", ...)
        target_dvt: Đơn vị kho trong Master List ("gram", "g", "kg", ...)

    Returns:
        Giá trị đã quy đổi sang target_dvt, hoặc None nếu không áp dụng được.
    """
    su = size_unit.lower().strip()
    tu = (target_dvt or "").lower().strip()

    # Bước 1: Chuẩn hoá size về ml
    if su in {"ml", "cc"}:
        size_ml = size_val
    elif su in {"l", "lít", "lít", "lit", "liter", "litre"}:
        size_ml = size_val * 1000
    else:
        return None  # Không phải volume → không áp dụng

    # Bước 2: Chuyển ml → target_dvt
    if tu in {"g", "gr", "gram", "gam"}:
        return size_ml          # 1ml = 1g
    elif tu in {"kg", "kgs"}:
        return size_ml / 1000   # 1ml = 0.001kg
    return None

def _convert_same_dimension(size_val: float, size_unit: str, target_dvt: str) -> float | None:
    su = size_unit.lower().strip()
    tu = (target_dvt or "").lower().strip()

    if su in _VOLUME_UNITS and tu in _VOLUME_UNITS:
        size_ml = size_val
        if su in {"l", "lít", "lit", "liter", "litre"}:
            size_ml = size_val * 1000
        
        if tu in {"ml", "cc"}:
            return size_ml
        elif tu in {"l", "lít", "lit", "liter", "litre"}:
            return size_ml / 1000

    elif su in _WEIGHT_UNITS and tu in _WEIGHT_UNITS:
        size_g = size_val
        if su in {"kg", "kgs", "ki-lô-gam"}:
            size_g = size_val * 1000
        elif su in {"mg"}:
            size_g = size_val / 1000
            
        if tu in {"g", "gr", "gram", "gam"}:
            return size_g
        elif tu in {"kg", "kgs"}:
            return size_g / 1000
        elif tu in {"mg"}:
            return size_g * 1000

    return None


def _try_size_in_name_conversion(
    raw_name: str,
    invoice_unit: str,
    master_record: dict,
    qty,
    unit_price,
):
    """
    Thử quy đổi đơn vị khi đơn vị hoá đơn không khớp DVT/DVT0 nhưng tên
    mặt hàng chứa kích thước.

    Trả về (qty_final, unit_price_final, unit_final, conversion_note, warning_note)
    hoặc (None, None, "", "", warning/empty) nếu không thể tự động quy đổi.

    Luồng xử lý:
    1. Trích xuất (size_val, size_unit) từ tên hàng
    2. Case A — same dimension: size_unit khớp DVT/DVT0 → quy đổi trực tiếp
       VD: "Phô mai 2kg" / gói → DVT=kg → 1 gói = 2kg
    3. Case B — dairy exception: volume vs weight nhưng là sữa tươi
       VD: "Sữa Meiji 950ml" / hộp → DVT=gram → 1 hộp = 950g (1ml=1g)
    4. Case C — khác dimension, không phải dairy → [UNIT_IN_NAME] warning
    """
    size_val, size_unit = _extract_size_from_name(raw_name)
    if not size_val or not size_unit:
        return None, None, "", "", ""

    master_dvt  = master_record.get("unit", "")
    master_dvt0 = master_record.get("dvt0", "")

    try:
        orig_qty = float(qty) if qty is not None else None
        orig_up  = float(unit_price) if unit_price is not None else None
    except (TypeError, ValueError):
        orig_qty = orig_up = None

    su = size_unit.lower()
    is_volume = su in _VOLUME_UNITS
    is_weight = su in _WEIGHT_UNITS

    # ── Phân loại dimension ───────────────────────────────────────────────────
    tu = (master_dvt or "").lower()
    size_dim = "volume" if is_volume else ("weight" if is_weight else "unknown")
    dvt_dim  = "volume" if tu in _VOLUME_UNITS else ("weight" if tu in _WEIGHT_UNITS else "other")

    # ── Case A: Cùng dimension (metric conversion) hoặc khớp đơn vị ───────────
    # Quy đổi thẳng về master_dvt nếu cùng hệ (volume/weight)
    if size_dim != "unknown" and size_dim == dvt_dim:
        converted_val = _convert_same_dimension(size_val, size_unit, master_dvt)
        if converted_val is not None and converted_val > 0 and orig_qty is not None:
            new_qty = round(orig_qty * converted_val, 3)
            new_up  = round(orig_up / converted_val, 4) if orig_up is not None else None
            note = (
                f"[Quy đổi DVT từ tên: {orig_qty} {invoice_unit} "
                f"\u00d7 {converted_val:g}{master_dvt} (từ {size_val:g}{size_unit}) = {new_qty:g} {master_dvt}]"
            )
            return new_qty, new_up, master_dvt, note, ""

    # Nếu không cùng dimension nhưng size_unit khớp DVT0 (ví dụ: hộp -> thùng, cần he_so0)
    if master_dvt0 and _match_unit(size_unit, master_dvt0) and orig_qty is not None and size_val > 0:
        new_qty = orig_qty * size_val
        new_up  = orig_up / size_val if orig_up is not None else None
        
        he_so0_raw = master_record.get("he_so0")
        if he_so0_raw:
            try:
                hs = float(he_so0_raw)
                if hs > 0:
                    final_qty = round(new_qty * hs, 3)
                    final_up = round(new_up / hs, 4) if new_up else None
                    note = (
                        f"[Quy đổi DVT từ tên & kho: {orig_qty} {invoice_unit} "
                        f"\u00d7 {size_val:g}{size_unit} \u00d7 {hs:g} = {final_qty:g} {master_dvt}]"
                    )
                    return final_qty, final_up, master_dvt, note, ""
            except ValueError:
                pass

    # ── Case B: dairy exception — volume vs weight ────────────────────────────
    # Lighthouse quy ước: 1ml = 1g cho sữa tươi
    # VD: "Sữa Meiji 950ml" / hộp → DVT=gram → 1 hộp = 950g
    if size_dim == "volume" and dvt_dim == "weight" and _is_dairy_liquid(raw_name):
        converted = _convert_volume_to_weight_dairy(size_val, size_unit, master_dvt)
        if converted is not None and converted > 0 and orig_qty is not None:
            new_qty = round(orig_qty * converted, 3)
            new_up  = round(orig_up / converted, 4) if orig_up is not None else None
            note = (
                f"[Quy đổi DVT sữa (1ml=1g): {orig_qty} {invoice_unit} "
                f"\u00d7 {converted:g}{master_dvt} = {new_qty:g} {master_dvt}]"
            )
            return new_qty, new_up, master_dvt, note, ""

    # ── Case C: khác dimension, không phải dairy → cảnh báo thủ công ─────────
    if size_dim != dvt_dim:
        warning = (
            f"[UNIT_IN_NAME: {size_val:g}{size_unit}/{invoice_unit} — "
            f"dimension {size_dim} vs DVT '{master_dvt}' ({dvt_dim}), "
            f"cần quy đổi thủ công]"
        )
        return None, None, "", "", warning

    return None, None, "", "", ""


def _is_high_risk(invoice_json: dict) -> bool:
    confidence_raw = invoice_json.get("document_info", {}).get("confidence_score", 1.0)
    
    # Ép kiểu an toàn, nếu là chữ (như "PRO_FALLBACK_HANDLED") thì mặc định là 1.0 (không coi là rủi ro về độ tin cậy)
    try:
        confidence = float(confidence_raw)
    except (ValueError, TypeError):
        confidence = 1.0

    supplier_code = invoice_json.get("supplier_info", {}).get("supplier_name_code", "")
    
    if confidence < HIGH_RISK_CONF_THRESHOLD:
        return True
    if not supplier_code or str(supplier_code).strip().lower() in ("", "mua lẻ", "mua le"):
        return True
    return False


def _is_mapping_risky(invoice_json: dict) -> bool:
    doc_info = invoice_json.get("document_info", {}) or {}
    confidence_raw = doc_info.get("confidence_score", 1.0)
    invoice_type = str(doc_info.get("invoice_type", "") or "").upper()
    recommended = str(doc_info.get("recommended_model", "") or "").upper()

    try:
        confidence = float(confidence_raw)
    except (ValueError, TypeError):
        # M-1: Đồng bộ với _is_high_risk: parse fail → confidence = 1.0 (không coi là rủi ro)
        confidence = 1.0

    if invoice_type == "HANDWRITTEN_NOTE":
        return True
    if recommended == "PRO":
        return True
    if confidence < HIGH_RISK_CONF_THRESHOLD:
        return True
    return False


def _row_has_math_error(item: dict) -> bool:
    try:
        qty        = float(item.get("quantity", 0) or 0)
        unit_price = float(item.get("unit_price", 0) or 0)
        total      = float(item.get("total_price", 0) or 0)
    except (ValueError, TypeError):
        return True # Dữ liệu không phải số -> coi như có lỗi

    if qty == 0 and total > 0:
        return True  # BUG-E5: qty=0 nhưng có thành tiền → OCR miss số lượng
    # FIX-1 (Case 4): qty có nhưng cả đơn giá lẫn thành tiền đều = 0 → thiếu giá tiền
    if qty > 0 and unit_price == 0 and total == 0:
        return True
    if qty == 0 or unit_price == 0:
        return False

    expected = round(qty * unit_price, 0)
    return abs(expected - total) > 5 # Nới lỏng sai số lên 5đ để tránh báo động giả do làm tròn


def _apply_fill_to_row(ws, row_num: int, fill: PatternFill, max_col: int = _TOTAL_COLUMNS):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.border = _THIN_BORDER
        cell.font = _FONT_DATA

def _apply_global_font(ws, size: int = 14):
    """Áp dụng cỡ chữ cho toàn bộ các ô có dữ liệu trong sheet, giới hạn vùng dùng được để tránh lag."""
    # Chỉ quét vùng có dữ liệu thực tế (max 1000 dòng để an toàn)
    limit_row = min(ws.max_row, 1000)
    limit_col = min(ws.max_column, 50)
    
    for row in ws.iter_rows(min_row=1, max_row=limit_row, min_col=1, max_col=limit_col):
        for cell in row:
            if cell.font:
                # Bảo tồn các thuộc tính quan trọng khác
                cell.font = Font(
                    name=cell.font.name or "Calibri",
                    size=size,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    vertAlign=cell.font.vertAlign,
                    underline=cell.font.underline,
                    strike=cell.font.strike,
                    color=cell.font.color
                )
            else:
                cell.font = Font(size=size)
        
        # Thiết lập độ cao dòng: 2 dòng đầu = 30 (header), còn lại = 18 (data)
        row_num = row[0].row
        ws.row_dimensions[row_num].height = 30 if row_num <= 2 else 18


def _auto_fit_columns(ws, max_rows: int = 1000, max_cols: int = 50):
    """
    Tự động điều chỉnh độ rộng cột dựa trên nội dung dài nhất.
    Sử dụng hệ số nhân 1.2 cho font size 14 để tránh bị cắt chữ.
    """
    limit_row = min(ws.max_row, max_rows)
    limit_col = min(ws.max_column, max_cols)

    for col_idx in range(1, limit_col + 1):
        max_len = 0
        for row in ws.iter_rows(min_row=1, max_row=limit_row, min_col=col_idx, max_col=col_idx):
            cell = row[0]
            if cell.value is None:
                continue
            
            # Nếu là công thức, không dùng độ dài chuỗi công thức để autofit (tránh bị quá rộng do VLOOKUP)
            val = cell.value
            if isinstance(val, str) and val.startswith("="):
                continue
                
            text = str(val) if val is not None else ""
            
            # Tính độ dài thực tế (đơn giản hóa)
            current_len = len(text)
            if current_len > max_len:
                max_len = current_len

        if max_len <= 0:
            continue

        col_letter = get_column_letter(col_idx)
        # Hệ số 1.2 cho font 14, cộng thêm padding 2
        target_width = min(70, max(10, (max_len + 2) * 1.15))
        ws.column_dimensions[col_letter].width = target_width

def _format_date(date_val: Any) -> str:
    """Chuẩn hoá ngày về dạng dd/mm/yyyy"""
    if not date_val:
        return ""
    if isinstance(date_val, (date, datetime)):
        return date_val.strftime("%d/%m/%Y")
    
    date_str = str(date_val).strip()
    # Thử các format phổ biến từ OCR
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(date_str, fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return date_str


def _get_item_department(raw_group: str) -> str:
    group = str(raw_group or "").strip().upper()
    for dept_key in ("RANG", "BANH", "BAR", "BEP"):
        if group.startswith(dept_key):
            return dept_key
    return ""


def _build_master_index(data_store: Any) -> Tuple[Dict[str, List[dict]], List[str]]:
    master_entries: Dict[str, List[dict]] = {}
    if data_store:
        if getattr(data_store, "items_by_name", None):
            for items in data_store.items_by_name.values():
                for item_idx, item in enumerate(items):  # C-2: use index for identity-safe check
                    display = item.get("name", "")
                    if not display:
                        continue
                    norm = _normalize_for_compare(display)
                    if not norm:
                        continue
                    master_entries.setdefault(norm, []).append(item)
        elif getattr(data_store, "items_dict", None):
            for item in data_store.items_dict.values():
                display = item.get("name", "")
                if not display:
                    continue
                norm = _normalize_for_compare(display)
                if not norm:
                    continue
                master_entries.setdefault(norm, []).append(item)

    return master_entries, list(master_entries.keys())


def _score_candidate(
    raw_name: str,
    candidate: dict,
    department_hint: str = "",
    unit_hint: str = "",
    price_hint: float | None = None,
    invoice_context: List[str] | None = None,
    supplier_products: str = "",
) -> Tuple[float, float, int]:
    cleaned = _clean_ocr_name(raw_name)
    raw_norm = _normalize_for_compare(cleaned)
    cand_name = candidate.get("name", "") or ""
    cand_norm = _normalize_for_compare(cand_name)

    name_score = 0.0
    if raw_norm and cand_norm:
        name_score = fuzz.token_set_ratio(raw_norm, cand_norm) / 100.0

    word_count = len(raw_norm.split()) if raw_norm else 0

    dept_score = None
    if department_hint:
        cand_dept = _get_item_department(candidate.get("group", ""))
        dept_score = 1.0 if cand_dept == department_hint else 0.0

    unit_score = None
    if unit_hint:
        cand_unit = candidate.get("unit", "")
        cand_dvt0 = candidate.get("dvt0", "")
        unit_score = 1.0 if (_match_unit(unit_hint, cand_unit) or _match_unit(unit_hint, cand_dvt0)) else 0.0

    price_score = None
    if price_hint is not None:
        price_ref = None
        if unit_hint and _match_unit(unit_hint, candidate.get("unit", "")):
            price_ref = candidate.get("gia_mua")
        elif unit_hint and _match_unit(unit_hint, candidate.get("dvt0", "")):
            price_ref = candidate.get("gia_mua0")
        else:
            price_ref = candidate.get("gia_mua") or candidate.get("gia_mua0")
        try:
            if price_ref is not None:
                price_ref = float(price_ref)
                diff = abs(float(price_hint) - price_ref) / price_ref if price_ref else 1.0
                if diff <= 0.30:
                    price_score = 1.0
                elif diff <= 0.50:
                    price_score = 0.5
                else:
                    price_score = 0.0
        except Exception:
            price_score = 0.0

    context_score = None
    if invoice_context:
        context_norm = _normalize_for_compare(" ".join([str(x) for x in invoice_context if x]))
        cand_tokens = set(cand_norm.split())
        context_tokens = set(context_norm.split())
        if cand_tokens and context_tokens and cand_tokens.intersection(context_tokens):
            context_score = 0.5
        else:
            context_score = 0.0

    supplier_score = None
    if supplier_products:
        supplier_score = 1.0 if cand_name and cand_name.lower() in supplier_products.lower() else 0.0

    weights = {
        "name": 0.45,
        "dept": 0.20,
        "unit": 0.15,
        "price": 0.12,
        "context": 0.05,
        "supplier": 0.03,
    }

    parts = [
        ("name", name_score),
        ("dept", dept_score),
        ("unit", unit_score),
        ("price", price_score),
        ("context", context_score),
        ("supplier", supplier_score),
    ]
    weight_sum = sum(weights[k] for k, v in parts if v is not None)
    if weight_sum <= 0:
        return 0.0, name_score, word_count
    score = sum(weights[k] * v for k, v in parts if v is not None) / weight_sum
    return score, name_score, word_count

# ---------------------------------------------------------------------------
# ITEM NAME MATCHING: Hybrid Local Fuzzy + LLM Fallback
# ---------------------------------------------------------------------------

import re
import unicodedata
from rapidfuzz import fuzz, process as rfprocess


def _clean_ocr_name(raw: str) -> str:
    """
    Tiền xử lý tên mặt hàng từ OCR trước khi so khớp.
    1. Loại bỏ phần tiếng Anh trong ngoặc đơn (case-insensitive)
    2. Loại bỏ đuôi text full-English (khi OCR scan cả tên EN không có ngoặc)
    3. Chuẩn hoá khoảng trắng thừa
    """
    text = str(raw).strip()
    # Bỏ phần ngoặc đơn (có thể chứa chữ hoa, thường, số, ký tự đặc biệt)
    text = re.sub(r'\([A-Za-z0-9\s\.\-\/\,\+\&]+\)', '', text)
    # Bỏ ngoặc đơn lẻ còn sót (ngoặc mở hoặc đóng rời)
    text = text.replace('(', '').replace(')', '')
    text = re.sub(r'\s+', ' ', text).strip()

    # Bỏ đuôi toàn ký tự Latin viết hoa (English noise)
    # CHỈ xóa nếu phần trước đó đã có chữ thường hoặc ký tự có dấu (dấu hiệu tiếng Việt)
    # Tránh xóa nhầm các tên thuần Anh như "COCA COLA CLASSIC"
    has_vietnamese = bool(re.search(r'[a-zà-ỹ]', text))
    if has_vietnamese:
        text = re.sub(r'\s+[A-Z][A-Z0-9\s\.\-\/\,\&]{4,}$', '', text)

    return text.strip()


def _normalize_for_compare(text: str) -> str:
    """
    Chuẩn hoá chuỗi để so sánh fuzzy:
    1. Sửa ký tự OCR lạ (ū→ư, ą→a...)
    2. Bỏ TOÀN BỘ dấu tiếng Việt (để "phô mai" khớp "phomai", "tương" khớp "turong")
    3. Bỏ ký tự đặc biệt, chỉ giữ chữ + số + khoảng trắng
    4. Chuẩn hoá đơn vị: gram→g, lít→l
    """
    # Map ký tự OCR hay nhầm
    replacements = {
        'ū': 'ư', 'ą': 'a', 'ę': 'e', 'ł': 'l', 'ō': 'ơ',
        'å': 'a', 'ń': 'n', 'ś': 's', 'ź': 'z',
    }
    result = text.lower()
    for bad, good in replacements.items():
        result = result.replace(bad, good)

    # Bỏ toàn bộ dấu tiếng Việt bằng Unicode decomposition
    # "phô mai" → "pho mai", "tương cà" → "tuong ca"
    nfkd = unicodedata.normalize('NFKD', result)
    result = ''.join(c for c in nfkd if not unicodedata.combining(c))

    # Đặc biệt: chữ "đ" không bị decompose → thay thủ công
    result = result.replace('đ', 'd')

    # Chuẩn hoá đơn vị trọng lượng
    result = re.sub(r'(\d+)\s*gram\b', r'\1g', result)
    result = re.sub(r'(\d+)\s*lit\b', r'\1l', result)
    result = re.sub(r'(\d+)\s*kg/tui\b', r'\1kg', result)

    # Chỉ giữ chữ + số + khoảng trắng + dấu chấm (cho trọng lượng như 3.3kg)
    result = re.sub(r'[^a-z0-9\s\.]', '', result)
    result = re.sub(r'\s+', ' ', result).strip()

    return result


# _fuzzy_match_items removed (H-3): dead code — pipeline uses _hybrid_map_items directly


def _llm_map_items(
    raw_names: List[str],
    data_store: Any,
    api_key: str,
    department_hint: str = "",
    supplier_code: str = "",
    supplier_products: str = "",
    invoice_context: List[str] = None,
    price_hints: Dict[str, float] = None,   # {raw_name -> unit_price từ hoá đơn}
    candidates_by_item: Dict[str, List[dict]] = None,
) -> Dict[str, str]:
    """
    Bước 2 (Fallback): Gọi LLM cho các item mà fuzzy match không giải quyết được.
    Sử dụng chuỗi suy luận 4 tầng ưu tiên:
      1. Bộ phận (department_hint)   — thu hẹp không gian tìm kiếm
      2. Nhà cung cấp (supplier_code)  — thu hẹp theo nhóm sản phẩm NCC
      3. Sản phẩm mẫu NCC            — prior nghiệp vụ mạnh nhất
      4. Ngữ cảnh chéo (invoice_context) — các mặt hàng khác trong cùng phiếu để phân giải mơ hồ
    """
    if not raw_names or not data_store or not api_key:
        return {}

    if not candidates_by_item:
        logging.error("🚨 [ExcelMapper] LLM Map: Không có candidate shortlist để gọi LLM.")
        return {}

    unique_names = list(set([str(n).strip() for n in raw_names if str(n).strip()]))
    if not unique_names:
        return {}

    # Build Candidate Catalog (shortlist only)
    seen = set()
    catalog_lines = []
    for item_list in candidates_by_item.values():
        for v in item_list:
            name = v.get("name", "")
            if not name:
                continue
            key = (
                name,
                v.get("unit", ""),
                v.get("dvt0", ""),
                v.get("gia_mua"),
                v.get("gia_mua0"),
                v.get("group", ""),
            )
            if key in seen:
                continue
            seen.add(key)
            dvt = v.get("unit", "")
            dvt0 = v.get("dvt0", "")
            gia_mua = v.get("gia_mua")
            gia_mua0 = v.get("gia_mua0")
            dept_label = _get_item_department(v.get("group", "")) or "N/A"
            price_parts = []
            if dvt and gia_mua is not None:
                price_parts.append(f"{dvt}:{int(gia_mua):,}")
            if dvt0 and gia_mua0 is not None:
                price_parts.append(f"{dvt0}:{int(gia_mua0):,}")
            price_str = " | ".join(price_parts)
            line = f"- {name} [Bộ phận: {dept_label}]"
            if dvt:
                line += f" [ĐVT: {dvt}]"
            if dvt0:
                line += f" [DVT0: {dvt0}]"
            if price_str:
                line += f" [Giá: {price_str}]"
            catalog_lines.append(line)
    master_entries = "\n".join(catalog_lines)

    # ------- Xây dựng chuỗi ưu tiên suy luận 3 tầng -------
    # Tầng 1 — BỘ PHẬN
    if department_hint:
        dept_block = (
            f"\n⎣ TẦNG 1 — BỘ PHẬN XÁC ĐỊNH:\n"
            f"Hoá đơn này thuộc bộ phận [{department_hint}].\n"
            f"Hãy ưu tiên các mặt hàng có [Bộ phận: {department_hint}] trong Master List.\n"
            f"Nếu từ điều kiện này không tìm được, mới mở rộng sang bộ phận khác."
        )
    else:
        dept_block = (
            "\n⎣ TẦNG 1 — BỘ PHẬN:\n"
            "Không có thông tin bộ phận. Hãy quan sát [Bộ phận: ...] trong Master List "
            "và chọn nhóm có tính nhất quán (Consistency) cao nhất dựa trên toàn bộ danh sách mặt hàng.\n"
        )

    # Tầng 2 — NHÀ CUNG CẤP
    if supplier_code:
        supplier_block = (
            f"\n⎣ TẦNG 2 — NHÀ CUNG CẤP:\n"
            f"Hoá đơn đến từ NCC mã [{supplier_code}].\n"
        )
    else:
        supplier_block = ""

    # Tầng 3 — SẢN PHẨM MẪU NCC
    if supplier_products:
        products_block = (
            f"\n⎣ TẦNG 3 — MẶT HÀNG ĐIỂN HÌNH CỦA NCC [{supplier_code}]:\n"
            f"{supplier_products}\n"
            f"Đây là ưu tiên nghiệp vụ: nếu tên OCR tương tự các sản phẩm này, ưu tiên match nhóm này trước."
        )
    else:
        products_block = ""

    # Tầng 4 — NGỮ CẢNH CHÉO (toàn bộ phiếu nhập, bao gồm cả các item đang cần map)
    if invoice_context and len(invoice_context) > 1:
        context_block = (
            f"\n⎣ TẦNG 4 — NGỮ CẢNH PHIẾU NHẬP (Cross-Invoice Context):\n"
            f"Toàn bộ phiếu này nhập các mặt hàng: {', '.join(invoice_context)}.\n"
            f"Hãy xét toàn cảnh: các mặt hàng xuất hiện cùng nhau gợi ý chủng loại vật tư của phiếu.\n"
            f"Ví dụ: phiếu có dâu tây, cam, thơm, bơ \u2192 đây là phiếu trái cây, nên 'bơ' = bơ avocado, không phải bơ đậu phụng."
        )
    else:
        context_block = ""

    # Tầng 5 — SO SÁNH GIÁ (chỉ khi có dữ liệu giá từ hoá đơn)
    price_hints = price_hints or {}
    has_price_data = any(v is not None for v in price_hints.values())
    if has_price_data:
        price_block = (
            "\n⎣ TẦNG 5 — SO SÁNH GIÁ (TÍN HIỆU PHỤ — chỉ dùng khi tên không phân giải được):\n"
            "Mỗi item trong Master List có [Giá: ĐVT:giá_tham_chiếu]. Quy tắc so sánh:\n"
            "- Nếu đơn vị trên hoá đơn khớp DVT chính (Gram, Kg, Hộp...) → so với giá DVT chính.\n"
            "- Nếu đơn vị trên hoá đơn khớp DVT0 (Kg, Thùng...) → so với giá DVT0.\n"
            "- Giá hoá đơn sai khác < 30% so với tham chiếu → ưu tiên item đó.\n"
            "- Giá hoá đơn khác biệt quá lớn (> 50%) → giảm điểm ưu tiên cho item đó.\n"
            "- PHỤ: Tên phân giải được rõ ràng → bỏ qua giá; Tên mơ hồ → dùng giá làm tín hiệu phân loại."
        )
    else:
        price_block = ""

    # Tầng 6 — TỪ ĐIỂN ALIAS (Do User xây dựng)
    if hasattr(data_store, 'aliases_dict') and data_store.aliases_dict:
        alias_lines = []
        for al, info in data_store.aliases_dict.items():
            code = info.get("code", "")
            if hasattr(data_store, 'items_by_code') and code in data_store.items_by_code:
                alias_lines.append(f"- '{al}' = {data_store.items_by_code[code]['name']}")
        if alias_lines:
            alias_block = (
                "\n⎣ TẦNG 6 — TỪ ĐIỂN VIẾT TẮT (ALIASES):\n"
                "Nếu tên OCR gần giống (bị lỗi typo/OCR nhẹ) với các từ viết tắt sau, hãy map về tên gốc tương ứng:\n"
                + "\n".join(alias_lines)
            )
        else:
            alias_block = ""
    else:
        alias_block = ""

    # Xây danh sách item cần map kèm shortlist
    items_lines = []
    for n in unique_names:
        cand_names = [c.get("name", "") for c in candidates_by_item.get(n, []) if c.get("name")]
        cand_text = "; ".join(cand_names[:8]) if cand_names else ""
        p = price_hints.get(n)
        price_text = f" [Giá hoá đơn: {int(p):,}]" if p is not None else ""
        items_lines.append(f"- {n}{price_text} | Candidates: {cand_text}")
    items_section = chr(10).join(items_lines)

    prompt = f"""Bạn là chuyên gia chuẩn hóa Danh Mục Vật Tư kho hàng F&B.
Dưới đây là danh sách tên vật tư đọc được từ hoá đơn (có thể viết tắt, sai dấu, lỗi OCR).
Hãy so khớp từng tên với MASTER LIST bằng cách suy luận theo THỨ TỰ ƯU TIÊN sau:
{dept_block}{supplier_block}{products_block}{context_block}{price_block}{alias_block}
⎣ QUY TẮC CHUNG:
- Phải đối chiếu với TOÀN BỘ Master List trước khi quyết định. Không dừng ở kết quả đầu tiên.
- Khi có nhiều kết quả tương đồng, dùng Tầng 4 (ngữ cảnh chéo) và Tầng 5 (giá) để phân giải.
- Chọn kết quả có xác suất đúng cao nhất (Best Probable Match).
- Tên OCR thường bị sai dấu tiếng Việt: "turong" = "tương", "pho mai" = "phomai"
- Bỏ qua phần tiếng Anh trong ngoặc đơn nếu có
- "125g" và "125gram" là CÙNG một sản phẩm; "3.3kg" có thể khớp item không ghi trọng lượng
- Trả về TÊN CHÍNH XÁC TỪ MASTER LIST (copy nguyên văn, KHÔNG kèm [Bộ phận])
- Nếu không chắc >= 60%, trả về ""

 CANDIDATE CATALOG (shortlist only):
 {master_entries}

 VẬT TƯ CẦN MAP (kèm giá + shortlist):
 {items_section}

Trả về JSON thuần (KHÔNG markdown):
{{
    "tên hoá đơn 1": "tên từ MASTER LIST",
    "tên hoá đơn 2": ""
}}
"""
    client = genai.Client(api_key=api_key)
    models_to_try = [
        (data_store.models.get("light_primary"), True),
        (data_store.models.get("light_fallback"), False),
    ]

    def _call_llm(call_prompt: str) -> dict:
        """Gọi LLM và trả về dict kết quả. Ném exception nếu thất bại hoàn toàn."""
        last_error = None
        for model_name, is_primary in models_to_try:
            if not model_name:
                continue
            try:
                global_rate_limiter.wait_if_needed("flash")
                config = None
                if data_store.should_use_minimal_thinking(model_name, is_primary=is_primary):
                    config = types.GenerateContentConfig(
                        thinking_config=types.ThinkingConfig(thinking_budget=0)
                    )
                response = client.models.generate_content(
                    model=model_name,
                    contents=[call_prompt],
                    config=config
                )
                raw = response.text.replace('```json', '').replace('```', '').strip()
                # Robust JSON parse: thử cắt từ { đầu đến } cuối để loại bỏ text thừa
                start = raw.find('{')
                end   = raw.rfind('}')
                if start != -1 and end != -1:
                    raw = raw[start:end + 1]
                return json.loads(raw)
            except Exception as e:
                last_error = e
                continue
        raise RuntimeError(f"LLM Mapping: all models failed — {last_error}")

    def _resolve_llm_keys(input_names: list, raw_result: dict) -> dict:
        """
        Khớp key từ LLM response về đúng input_name gốc.
        LLM đôi khi trả key hơi khác (thêm khoảng trắng, capitalize, bỏ dấu)
        → dùng fuzzy match trên key để tránh silent miss.
        """
        resolved = {}
        used_keys = set()
        for name in input_names:
            # 1. Khớp chính xác (fast path)
            if name in raw_result:
                resolved[name] = raw_result[name]
                used_keys.add(name)
                continue
            # 2. Khớp case-insensitive + strip
            name_lower = name.strip().lower()
            exact_ci = next(
                (k for k in raw_result if k.strip().lower() == name_lower and k not in used_keys),
                None
            )
            if exact_ci:
                resolved[name] = raw_result[exact_ci]
                used_keys.add(exact_ci)
                continue
            # 3. Fuzzy fallback trên key (chỉ dùng khi ≥ 88 để tránh nhầm)
            name_norm = _normalize_for_compare(name)
            best_key, best_score = None, 0
            for k in raw_result:
                if k in used_keys:
                    continue
                score = fuzz.ratio(name_norm, _normalize_for_compare(k))
                if score > best_score:
                    best_score, best_key = score, k
            if best_key and best_score >= 88:
                resolved[name] = raw_result[best_key]
                used_keys.add(best_key)
            else:
                resolved[name] = ""  # Thực sự không có trong response
        return resolved

    try:
        llm_result = _call_llm(prompt)
        resolved   = _resolve_llm_keys(unique_names, llm_result)

        # Retry 1 lần cho các item vẫn còn trống sau lần đầu
        still_missing = [n for n, v in resolved.items() if not v]
        if still_missing and len(still_missing) < len(unique_names):
            # BUG-E6 fix: Xây prompt retry RIÊNG BIỆT (không nối vào prompt gốc
            # vì prompt gốc đã có section VẬT TƯ CẦN MAP → LLM confused 2 danh sách)
            retry_items_text = chr(10).join(f"- {n}" for n in still_missing)
            retry_prompt = f"""Bạn là chuyên gia chuẩn hóa Danh Mục Vật Tư kho hàng F&B.
Lần trước bạn đã bỏ sót {len(still_missing)} vật tư. Hãy cố gắng map lại.
{dept_block}{supplier_block}{products_block}{context_block}{alias_block}
⎣ QUY TẮC: Đối chiếu TOÀN BỘ Master List. Trả TÊN CHÍNH XÁC. Nếu không chắc >= 60%, trả "".

MASTER LIST:
{master_entries}

VẬT TƯ CẦN MAP:
{retry_items_text}

Trả về JSON thuần (KHÔNG markdown):
{{{{
    "tên hoá đơn": "tên từ MASTER LIST"
}}}}
"""
            try:
                retry_result   = _call_llm(retry_prompt)
                retry_resolved = _resolve_llm_keys(still_missing, retry_result)
                for name, val in retry_resolved.items():
                    if val and not resolved.get(name):
                        resolved[name] = val
            except Exception:
                pass  # Retry thất bại — giữ kết quả từ lần đầu

        return resolved

    except Exception as e:
        print(f"LLM Mapping Error (All models failed): {e}")
        return {}


def _infer_department(item_codes: List[str]) -> str:
    """
    Suy ra bộ phận của hoá đơn dựa trên tiền tố của mã vật tư trong items_dict.
    """
    counts = {"BAR": 0, "BEP": 0, "BANH": 0, "RANG": 0}  # BUG-01 fix: thêm RANG
    for code in item_codes:
        c = str(code).upper()
        if c.startswith(("BAR", "RUOU", "CCDC-BAR", "CCDCBAR")):
            counts["BAR"] += 1
        elif c.startswith(("BANH", "CCDC-BANH", "CCDCBANH")):
            counts["BANH"] += 1
        elif c.startswith(("BEP", "CCDC-BEP", "CCDCBEP")):
            counts["BEP"] += 1
        elif c.startswith(("RANG",)):  # BUG-01 fix
            counts["RANG"] += 1
    dominant = max(counts, key=counts.get)
    return dominant if counts[dominant] > 0 else ""


def _resolve_candidate_record(
    raw_name: str,
    mapped_name: str,
    data_store: Any,
    department_hint: str = "",
    unit_hint: str = "",
    price_hint: float | None = None,
    invoice_context: List[str] | None = None,
    supplier_products: str = "",
) -> dict:
    if not mapped_name or not data_store:
        return {}

    candidates = []
    if getattr(data_store, "items_by_name", None):
        candidates = data_store.items_by_name.get(mapped_name.lower(), [])

    if not candidates and getattr(data_store, "items_dict", None):
        fallback = data_store.items_dict.get(mapped_name.lower())
        if fallback:
            return fallback

    if not candidates:
        return {}

    best = None
    best_score = -1.0
    for cand in candidates:
        score, _name_score, _word_count = _score_candidate(
            raw_name=raw_name,
            candidate=cand,
            department_hint=department_hint,
            unit_hint=unit_hint,
            price_hint=price_hint,
            invoice_context=invoice_context,
            supplier_products=supplier_products,
        )
        if score > best_score:
            best_score = score
            best = cand
    return best or {}


def _hybrid_map_items(
    raw_names: List[str],
    data_store: Any,
    api_key: str,
    department_hint: str = "",
    supplier_code: str = "",
    supplier_products: str = "",
    status_callback=None,
    price_hints: Dict[str, float] = None,  # {raw_name -> unit_price từ hoá đơn}
    unit_hints: Dict[str, str] = None,     # {raw_name -> unit từ hoá đơn}
    is_risky: bool = False,
) -> Tuple[Dict[str, str], Dict[str, float]]:
    """
    Pipeline so khớp 2 tầng:
    Tầng 1 — Local scoring (fuzzy + dept + unit + price + context)
    Tầng 2 — LLM fallback (chỉ cho item miss) với shortlist candidates
    """
    if not raw_names:
        return {}, {}

    unique_names = list(set([str(n).strip() for n in raw_names if str(n).strip()]))
    unit_hints = unit_hints or {}
    price_hints = price_hints or {}

    master_entries, master_norm_list = _build_master_index(data_store)
    if not master_entries:
        logging.error("🚨 [ExcelMapper] Master List đang rỗng, không thể map.")
        return {}, {}

    matched: Dict[str, str] = {}
    unmatched: List[str] = []
    score_map: Dict[str, float] = {}
    candidates_by_item: Dict[str, List[dict]] = {}

    map_threshold = 0.82 if is_risky else 0.78
    margin_threshold = 0.15 if is_risky else 0.12

    for raw in unique_names:
        cleaned = _clean_ocr_name(raw)
        
        # --- BƯỚC 0: KIỂM TRA TỪ ĐIỂN ALIAS ---
        alias_key = cleaned.strip().lower()
        if hasattr(data_store, 'aliases_dict') and alias_key in data_store.aliases_dict:
            alias_info = data_store.aliases_dict[alias_key]
            code = alias_info.get("code", "")
            # Lookup the code to get the standard name
            if hasattr(data_store, 'items_by_code') and code in data_store.items_by_code:
                matched[raw] = data_store.items_by_code[code]["name"]
                score_map[raw] = 1.0
                logging.info(f"[Alias] '{raw}' -> Code: '{code}' ({matched[raw]})")
                continue

        query = _normalize_for_compare(cleaned)
        if not query:
            unmatched.append(raw)
            score_map[raw] = 0.0
            continue

        matches = rfprocess.extract(
            query, master_norm_list,
            scorer=fuzz.token_set_ratio,
            score_cutoff=60,
            limit=25
        )

        candidates = []
        for matched_norm, _score, _ in matches:
            candidates.extend(master_entries.get(matched_norm, []))

        if not candidates:
            unmatched.append(raw)
            score_map[raw] = 0.0
            continue

        scored = []
        for cand in candidates:
            score, name_score, word_count = _score_candidate(
                raw_name=raw,
                candidate=cand,
                department_hint=department_hint,
                unit_hint=unit_hints.get(raw, ""),
                price_hint=price_hints.get(raw),
                invoice_context=unique_names,
                supplier_products=supplier_products,
            )
            scored.append((score, name_score, word_count, cand))

        scored.sort(key=lambda x: x[0], reverse=True)
        best_score, best_name_score, best_word_count, best_cand = scored[0]
        second_score = scored[1][0] if len(scored) > 1 else 0.0

        candidates_by_item[raw] = [s[3] for s in scored[:8]]
        score_map[raw] = best_score

        if best_word_count <= 2 and best_name_score < 0.88:
            unmatched.append(raw)
            continue

        if best_score >= map_threshold and (best_score - second_score) >= margin_threshold:
            matched[raw] = best_cand.get("name", "")
        else:
            unmatched.append(raw)

    msg = f"[ExcelMapper] Local map: {len(matched)}/{len(unique_names)} khớp. Chưa khớp: {unmatched if unmatched else 'không có'}"
    logging.info(msg)
    if status_callback:
        status_callback(msg)

    # Tầng 2: LLM chỉ cho nhóm chưa khớp, gửi kèm shortlist
    llm_results = {}
    if unmatched and api_key:
        msg_start = f"[ExcelMapper] >>> {len(unmatched)} vat tu chua khop. Khoi dong LLM Fallback..."
        logging.info(msg_start)
        if status_callback:
            status_callback(msg_start)

        llm_results = _llm_map_items(
            unmatched, data_store, api_key,
            department_hint=department_hint,
            supplier_code=supplier_code,
            supplier_products=supplier_products,
            invoice_context=unique_names,
            price_hints=price_hints,
            candidates_by_item={k: v for k, v in candidates_by_item.items() if k in unmatched},
        )
        matched_by_llm = sum(1 for v in llm_results.values() if v)
        still_empty = sum(1 for v in llm_results.values() if not v)
        for item_name, mapped_to in llm_results.items():
            result_str = mapped_to if mapped_to else "[Không tìm thấy]"
            item_msg = f"[ExcelMapper] LLM: '{item_name}' => '{result_str}'"
            logging.info(item_msg)
            if status_callback:
                status_callback(item_msg)
        summary_msg = (
            f"[ExcelMapper] LLM kết quả: {matched_by_llm}/{len(unmatched)} map được"
            + (f" | {still_empty} vẫn trống (tô vàng)" if still_empty else "")
        )
        logging.info(summary_msg)
        if status_callback:
            status_callback(summary_msg)
    elif unmatched:
        msg_warn = f"[ExcelMapper] {len(unmatched)} vat tu chua khop nhung khong co API Key de chay LLM Fallback."
        logging.warning(msg_warn)
        if status_callback:
            status_callback(msg_warn)

    final = {}
    for name in unique_names:
        if name in matched:
            final[name] = matched[name]
        elif name in llm_results and llm_results[name]:
            final[name] = llm_results[name]
            score_map[name] = max(score_map.get(name, 0.0), 0.70)
        else:
            final[name] = ""

    return final, score_map


# ---------------------------------------------------------------------------
# MAIN EXPORT FUNCTION
# ---------------------------------------------------------------------------

def append_invoices_to_excel(invoice_results: List[Dict[str, Any]], data_store=None, api_key: str = None, output_dir: str = None, status_callback=None) -> str:
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    # Save to caller-specified dir (e.g. the DONE folder next to input images),
    # falling back to _BASE_DIR/DONE for backward compatibility.
    target_dir  = output_dir if output_dir else os.path.join(_ROOT_DIR, "DONE")
    output_file = os.path.join(target_dir, f"Lighthouse_NHAP_HANG_{timestamp}.xlsx")
    
    # Ensure target directory exists
    os.makedirs(target_dir, exist_ok=True)
    
    # Copy Template instead of creating blank
    if not os.path.exists(_TEMPLATE_FILE):
        raise FileNotFoundError(f"Không tìm thấy template mẫu: {_TEMPLATE_FILE}")
    shutil.copy(_TEMPLATE_FILE, output_file)

    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    
    # (Đã dời _apply_global_font xuống cuối để tối ưu tốc độ và độ chính xác)
    
    # Initialize Notes Header Column AA (Col 27)
    header_cell = ws.cell(row=2, column=27, value="NOTES (Hệ thống OCR)")
    header_cell.font = _FONT_HEADER
    header_cell.fill = _FILL_HEADER
    header_cell.alignment = Alignment(horizontal="center", vertical="center")
    header_cell.border = _THIN_BORDER
    comment = Comment(_NOTES_CELL_COMMENT, "Lighthouse OCR")
    comment.width = 450   # Đủ rộng để không bị cắt chữ
    comment.height = 200  # Đủ cao cho toàn bộ nội dung
    header_cell.comment = comment
    ws.column_dimensions[openpyxl.utils.get_column_letter(27)].width = 40

    current_row = _DATA_START_ROW
    
    # Auto-find the first empty data row (if template was previously populated)
    while ws.cell(row=current_row, column=1).value is not None or ws.cell(row=current_row, column=2).value is not None:
        current_row += 1

    # --- [HARDENING] Tạo Sheet ẩn & Dropdown (bọc try-except để tránh crash toàn hệ thống) ---
    master_sheet_name = "_HIDDEN_MASTER_DATA"
    last_m_row = 1
    try:
        if master_sheet_name in wb.sheetnames:
            del wb[master_sheet_name]
        ms = wb.create_sheet(master_sheet_name)
        ms.sheet_state = 'hidden'
        
        # 1. Danh mục Vật tư (Cột A: Tên, B: Mã, C: ĐVT)
        all_master_items = []
        if data_store and data_store.items_by_name:
            for name_key, records in data_store.items_by_name.items():
                all_master_items.extend(records)
        all_master_items.sort(key=lambda x: str(x.get('name', '')).lower())

        name_counts = {}
        for item in all_master_items:
            n = item.get('name', '').lower()
            name_counts[n] = name_counts.get(n, 0) + 1

        for idx, item_v in enumerate(all_master_items, 1):
            disp_name = item_v.get('name', '')
            if name_counts.get(disp_name.lower(), 0) > 1:
                code_suffix = item_v.get('code', 'N/A')
                disp_name = f'{disp_name} ({code_suffix})'
            ms.cell(row=idx, column=1, value=disp_name)
            ms.cell(row=idx, column=2, value=item_v.get('code', ''))
            ms.cell(row=idx, column=3, value=item_v.get('unit', ''))
        last_m_row = len(all_master_items) if all_master_items else 1

        # 2. Danh mục Đối tượng/NCC (Cột E)
        all_suppliers = sorted([str(v) for v in data_store.suppliers_dict.values()]) if data_store else []
        for idx, s_name in enumerate(all_suppliers, 1):
            ms.cell(row=idx, column=5, value=s_name)
        last_s_row = len(all_suppliers) if all_suppliers else 1

        # 3. Danh mục Kho (Cột G)
        all_khos = sorted(list(set([str(v) for v in data_store.kho_dict.values()]))) if data_store else []
        for idx, k_name in enumerate(all_khos, 1):
            ms.cell(row=idx, column=7, value=k_name)
        last_k_row = len(all_khos) if all_khos else 1

        # 4. Danh mục Bộ phận (Cột I)
        all_depts = ["BAR", "BEP", "BANH", "RANG"]
        for idx, d_name in enumerate(all_depts, 1):
            ms.cell(row=idx, column=9, value=d_name)
        last_d_row = len(all_depts)

        # --- Tạo các đối tượng Validation (Non-Strict: showErrorMessage=False) ---
        def _create_dv(formula, prompt, title):
            return DataValidation(
                type="list", formula1=formula, allow_blank=True,
                showErrorMessage=False, showInputMessage=True,
                promptTitle=title, prompt=prompt
            )

        dv_item = _create_dv(f"='{master_sheet_name}'!$A$1:$A${last_m_row}", 
                           "Chọn từ danh mục hoặc tự nhập tên mới. Mã và ĐVT sẽ tự cập nhật.", "Diễn giải (Vật tư)")
        dv_supplier = _create_dv(f"='{master_sheet_name}'!$E$1:$E${last_s_row}", 
                               "Chọn NCC từ danh mục hoặc tự nhập mới.", "Đối tượng (NCC)")
        dv_kho = _create_dv(f"='{master_sheet_name}'!$G$1:$G${last_k_row}", 
                          "Chọn kho nhập phù hợp.", "Kho nhập")
        dv_dept = _create_dv(f"='{master_sheet_name}'!$I$1:$I${last_d_row}", 
                           "Chọn bộ phận hạch toán.", "Bộ phận")

        ws.add_data_validation(dv_item)
        ws.add_data_validation(dv_supplier)
        ws.add_data_validation(dv_kho)
        ws.add_data_validation(dv_dept)

        # Áp dụng cho dải 1000 dòng
        dv_item.add(f"I{current_row}:I{current_row + 1000}")
        dv_supplier.add(f"H{current_row}:H{current_row + 1000}")
        dv_kho.add(f"K{current_row}:K{current_row + 1000}")
        dv_dept.add(f"F{current_row}:F{current_row + 1000}")

    except Exception as e:
        logging.error(f"⚠️ Lỗi khởi tạo Data Validation: {e}")

    items_db = data_store.items_dict if data_store else {}

    # Chi phí workbook (lazy init — chỉ tạo khi có item unmapped + NCC đã rõ)
    chiphi_wb = None
    chiphi_ws = None
    chiphi_output = os.path.join(target_dir, f"Lighthouse_CHI_PHI_{timestamp}.xlsx")
    chiphi_current_row = 3   # Mặc định bắt đầu từ dòng 3, sẽ được update khi load file
    chiphi_count = 0
    _CHIPHI_COLS = 15        # Template Chi phí có 15 cột (thêm cột Notes)
    
    # Map từng invoice riêng biệt để giữ nguyên department context
    per_invoice_maps: Dict[int, Dict[str, str]] = {}
    per_invoice_scores: Dict[int, Dict[str, float]] = {}
    per_invoice_context: Dict[int, List[str]] = {}
    per_invoice_supplier_products: Dict[int, str] = {}
    for inv_idx, inv in enumerate(invoice_results):
        raw_names_inv = [
            item.get("product_name") for item in inv.get("items", [])
            if item.get("product_name")
        ]
        if not raw_names_inv:
            per_invoice_maps[inv_idx] = {}
            per_invoice_scores[inv_idx] = {}
            per_invoice_context[inv_idx] = []
            per_invoice_supplier_products[inv_idx] = ""
            continue

        # ✅ Đọc bộ phận trực tiếp từ JSON (OCR đã trích xuất từ tiêu đề phiếu)
        dept_hint = str(inv.get("transaction_info", {}).get("department", "") or "").strip().upper()
        if dept_hint not in ("BEP", "BAR", "BANH", "RANG"):
            dept_hint = ""

        # ✅ Lấy thông tin Nhà Cung Cấp cho Tầng 2 và Tầng 3 của LLM context
        inv_supplier_code = str(inv.get("supplier_info", {}).get("supplier_name_code", "") or "").strip()
        # Tìm sản phẩm mẫu của NCC trong suppliers_context_str
        # Format: "TCAY=Châu Thủy (Sản phẩm: Bơ, Cam, Dâu tây) | BACON=..."
        inv_supplier_products = ""
        if inv_supplier_code and data_store and data_store.suppliers_context_str:
            # Tìm đoạn bắt đầu bằng "CODE=" cho đến "|" tiếp theo hoặc cuối chuỗi
            pattern = rf'(?:^|\| ){re.escape(inv_supplier_code)}=([^|]+)'
            m_sup = re.search(pattern, data_store.suppliers_context_str)
            if m_sup:
                # Dùng group(1) lấy đúng capture group (không kèm prefix "CODE=")
                inv_supplier_products = m_sup.group(1).strip()

        # Tầng 5: Xây price_hints {tên_thô -> unit_price} để LLM so sánh giá
        inv_price_hints: Dict[str, float] = {}
        inv_unit_hints: Dict[str, str] = {}
        for _item in inv.get("items", []):
            _pname  = _item.get("product_name")
            _uprice = _item.get("unit_price")
            _unit   = _item.get("unit")
            if _pname and _uprice is not None:
                try:
                    inv_price_hints[str(_pname).strip()] = float(_uprice)
                except (ValueError, TypeError):
                    pass
            if _pname and _unit:
                inv_unit_hints[str(_pname).strip()] = str(_unit).strip()

        inv_maps, inv_scores = _hybrid_map_items(
            raw_names_inv, data_store, api_key,
            department_hint=dept_hint,
            supplier_code=inv_supplier_code,
            supplier_products=inv_supplier_products,
            status_callback=status_callback,
            price_hints=inv_price_hints,
            unit_hints=inv_unit_hints,
            is_risky=_is_mapping_risky(inv),
        )
        per_invoice_maps[inv_idx] = inv_maps
        per_invoice_scores[inv_idx] = inv_scores
        per_invoice_context[inv_idx] = raw_names_inv
        per_invoice_supplier_products[inv_idx] = inv_supplier_products

    for invoice_idx, invoice_json in enumerate(invoice_results):
        # Fallback raw if code is missing
        supplier_code = invoice_json.get("supplier_info", {}).get("supplier_name_code", "")
        if not supplier_code:
            supplier_code = invoice_json.get("supplier_info", {}).get("supplier_name_raw", "")

        supplier_known = bool(supplier_code) and str(supplier_code).strip().lower() not in ("", "mua lẻ", "mua le")
        mapping_risky = _is_mapping_risky(invoice_json)
            
        items    = invoice_json.get("items", [])
        txn_info = invoice_json.get("transaction_info", {})
        
        # ✅ Đọc bộ phận từ JSON (nguon chính xác nhat: trích xuất từ tiêu đề phiếu bởi OCR)
        invoice_dept = str(txn_info.get("department", "") or "").strip().upper()
        if invoice_dept not in ("BEP", "BAR", "BANH", "RANG"):
            invoice_dept = None
        
        voucher_number = None  # C-3: lazy — only allocated when first item is written
        def _get_voucher():
            nonlocal voucher_number
            if voucher_number is None:
                voucher_number = get_next_voucher_number()
            return voucher_number
        high_risk      = _is_high_risk(invoice_json)
        base_fill      = _get_fill_for_invoice(invoice_idx, high_risk)

        # Log cảnh báo nếu thiếu dept hoặc NCC
        if status_callback:
            if not invoice_dept:
                status_callback(f"⚠️ [{_get_voucher()}] Không xác định được Bộ phận trên hoá đơn")
            ncc_code_raw = str(invoice_json.get("supplier_info", {}).get("supplier_name_code", "") or "").strip()
            if not ncc_code_raw or ncc_code_raw.lower() in ("mua lẻ", "mua le"):
                status_callback(f"⚠️ [{_get_voucher()}] NCC trống/mua lẻ — giữ item unmapped trong PNMH")

        for item_idx, item in enumerate(items):  # C-2: use index for identity-safe check
            # H-2: math_error evaluated AFTER unit conversion (see below)
            raw_product_name = item.get("product_name", "") or ""
            invoice_unit_raw = str(item.get("unit", "") or "").strip()

            # --- ÁP DỤNG ĐƠN VỊ TÍNH LÓNG ---
            alias_key = _clean_ocr_name(raw_product_name).strip().lower()
            if hasattr(data_store, 'aliases_dict') and alias_key in data_store.aliases_dict:
                alias_unit = data_store.aliases_dict[alias_key].get("unit", "")
                if not invoice_unit_raw and alias_unit:
                    invoice_unit_raw = alias_unit
                    existing_notes = str(item.get("notes", "") or "")
                    item["notes"] = f"{existing_notes}, [Tự động: Áp dụng ĐVT lóng '{alias_unit}' từ Từ điển]".strip(", ")
                    logging.info(f"[Alias Unit] '{raw_product_name}' thiếu ĐVT -> tự động gán ĐVT lóng '{alias_unit}'")

            # LLM trả về display_name (tên hiển thị chuẩn từ Master List), scoped per invoice
            mapped_display_name = per_invoice_maps.get(invoice_idx, {}).get(str(raw_product_name).strip(), "").strip()
            map_score = per_invoice_scores.get(invoice_idx, {}).get(str(raw_product_name).strip(), 0.0)

            # Tra ngược candidates theo tên đã map (item-centric, không overwrite)
            m = {}
            if mapped_display_name:
                m = _resolve_candidate_record(
                    raw_name=str(raw_product_name).strip(),
                    mapped_name=mapped_display_name,
                    data_store=data_store,
                    department_hint=invoice_dept or "",
                    unit_hint=invoice_unit_raw,
                    price_hint=item.get("unit_price"),
                    invoice_context=per_invoice_context.get(invoice_idx, []),
                    supplier_products=per_invoice_supplier_products.get(invoice_idx, ""),
                )
                
                # --- BƠM HỆ SỐ LÓNG TỪ TỪ ĐIỂN VÀO LOGIC QUY ĐỔI ---
                alias_key = _clean_ocr_name(raw_product_name).strip().lower()
                if hasattr(data_store, 'aliases_dict') and alias_key in data_store.aliases_dict:
                    alias_info = data_store.aliases_dict[alias_key]
                    
                    # Tìm đơn vị lóng khớp với hóa đơn trong danh sách đa đơn vị
                    target_unit_info = None
                    if invoice_unit_raw:
                        for u_info in alias_info.get("units", []):
                            if _match_unit(invoice_unit_raw, u_info["unit"]):
                                target_unit_info = u_info
                                break
                    
                    if target_unit_info:
                        alias_factor = target_unit_info.get("factor")
                        alias_unit = target_unit_info.get("unit", "")
                        
                        if alias_factor is not None:
                            m["dvt0"] = alias_unit
                            m["he_so0"] = alias_factor
                            logging.info(f"[Alias Factor] Bơm hệ số lóng {alias_factor} cho đơn vị '{alias_unit}'")

            mapped_code = m.get("code", "")
            mapped_unit = m.get("unit", "")

            # Nếu không map được → dùng tên thô từ hoá đơn (không bỏ trống)
            item_unmapped = not bool(m)
            if item_unmapped:
                final_name = str(raw_product_name).strip() or "[Không đọc được tên]"
            else:
                final_name = m.get("name", raw_product_name)

            # Chuẩn bị notes: thêm cảnh báo nếu item unmapped
            existing_notes = str(item.get("notes", "") or "")

            # FIX-4B: Lan truyền cảnh báo lệch tổng từ Calculator lên notes item đầu tiên
            discrepancy_warning = str(invoice_json.get("totals", {}).get("total_discrepancy_warning", "") or "")

            if item_unmapped and raw_product_name:
                unmapped_warning = f"[Chưa khớp Master List: {raw_product_name}]"
                # FIX-5: Giải thích rõ lý do khi item ở lại PNMH do NCC trống
                if not supplier_known:
                    unmapped_warning += " [Giữ PNMH: NCC chưa xác định — không thể quyết định loại chi phí]"
                notes_out = f"{existing_notes}, {unmapped_warning}".strip(", ")
            elif item_unmapped:
                notes_out = f"{existing_notes}, [Tên vật tư trống]".strip(", ")
            else:
                notes_out = existing_notes

            # Gắn cảnh báo lệch tổng vào item đầu tiên của invoice (items.index == 0)
            if discrepancy_warning and item_idx == 0:  # C-2: index check, not identity
                notes_out = f"{notes_out}, {discrepancy_warning}".strip(", ") if notes_out else discrepancy_warning

            # --- Business Logic: BỘ PHẬN ---
            # Nguồn ưu tiên: Lấy từ JSON (OCR đã đọc trên phiếu)
            # Fallback: Suy ra từ tiền tố mã vật tư (nếu OCR không bắt được)
            bo_phan = invoice_dept
            if not bo_phan:
                upper_code = str(mapped_code).upper().strip() if mapped_code else ""
                if upper_code.startswith(("BAR", "RUOU", "CCDCBAR")):
                    bo_phan = "BAR"
                elif upper_code.startswith("BEP"):
                    bo_phan = "BEP"
                elif upper_code.startswith(("BANH", "CCDCBANH")):
                    bo_phan = "BANH"
                elif upper_code.startswith("RANG"):
                    bo_phan = "RANG"

            # --- Quy đổi đơn vị: Nếu hoá đơn dùng Dvt0 thay vì Dvt kho ---
            # Ví dụ: Hoá đơn ghi "2 Kg", kho lưu bằng "Gram", He_So0=1000
            #   → Số lượng quy đổi = 2 × 1000 = 2000 Gram
            #   → Đơn giá quy đổi = đơn_giá_invoice / 1000 = giá mỗi Gram
            qty_final        = item.get("quantity", None)
            unit_price_final = item.get("unit_price", None)
            total_price_final= item.get("total_price", None)
            unit_final       = mapped_unit if mapped_unit else invoice_unit_raw
            conversion_note  = ""

            if m and invoice_unit_raw:
                master_dvt0 = m.get("dvt0", "")
                he_so0      = m.get("he_so0")   # float hoặc None
                master_dvt  = m.get("unit", "") # ĐVT kho (đích)

                if master_dvt0 and he_so0 and _match_unit(invoice_unit_raw, master_dvt0):
                    try:
                        he_so0_f = float(he_so0)
                        if he_so0_f > 0:
                            orig_qty = float(qty_final) if qty_final is not None else None
                            orig_up  = float(unit_price_final) if unit_price_final is not None else None

                            if orig_qty is not None:
                                qty_final = round(orig_qty * he_so0_f, 3)
                            if orig_up is not None:
                                unit_price_final = round(orig_up / he_so0_f, 4)
                            # total_price giữ nguyên (= số tiền thực trả, không đổi)

                            unit_final = master_dvt  # Ghi đVT kho vào Excel
                            conversion_note = (
                                f"[Quy đổi đVT: {orig_qty} {invoice_unit_raw} × {he_so0_f:g} "
                                f"= {qty_final:g} {master_dvt}]"
                            )
                    except (TypeError, ValueError):
                        pass  # Giữ nguyên nếu quy đổi lỗi

            # Gắn ghi chú quy đổi vào notes
            if conversion_note:
                notes_out = f"{notes_out}, {conversion_note}".strip(", ") if notes_out else conversion_note

            # Size-in-Name Fallback: khi DVT/DVT0 không khớp với đơn vị trên phiếu
            # nhưng tên mặt hàng chứa kích thước (VD: "Phô mai 2kg" unit="gói")
            if not conversion_note and m and invoice_unit_raw:
                sn_qty, sn_up, sn_unit, sn_note, sn_warning = _try_size_in_name_conversion(
                    raw_name=raw_product_name,
                    invoice_unit=invoice_unit_raw,
                    master_record=m,
                    qty=qty_final,
                    unit_price=unit_price_final,
                )
                if sn_note:
                    # Case A: tự động quy đổi được (VD: "gói 2kg" → kg)
                    qty_final = sn_qty
                    unit_price_final = sn_up
                    unit_final = sn_unit
                    notes_out = f"{notes_out}, {sn_note}".strip(", ") if notes_out else sn_note
                    logging.info(f"[ExcelMapper] Size-in-name conversion: {sn_note}")
                elif sn_warning:
                    # Case B: khác dimension (VD: hộp 950ml vs gram) — cảnh báo thủ công
                    notes_out = f"{notes_out}, {sn_warning}".strip(", ") if notes_out else sn_warning
                    logging.warning(f"[ExcelMapper] Unit-in-name mismatch: {sn_warning}")

            # H-2: Đánh giá lỗi toán học SAU quy đổi đơn vị để phát hiện sai số post-conversion
            post_conv_item = dict(item)
            post_conv_item["quantity"] = qty_final
            post_conv_item["unit_price"] = unit_price_final
            post_conv_item["total_price"] = total_price_final
            math_error = _row_has_math_error(post_conv_item)

            # --- ROUTING DECISION ---
            # unmapped + NCC đã rõ (not high_risk) → item không có trong danh mục → Chi phí
            # unmapped + NCC trống (high_risk)     → có thể do thiếu context → giữ PNMH + vàng
            route_to_chiphi = (
                item_unmapped
                and supplier_known
                and not mapping_risky
                and map_score < 0.40
            )

            if route_to_chiphi:
                # Lazy init Chi phí workbook
                if chiphi_wb is None:
                    if not os.path.exists(_CHIPHI_TEMPLATE_FILE):
                        route_to_chiphi = False  # Không có template → fallback về PNMH
                    else:
                        shutil.copy(_CHIPHI_TEMPLATE_FILE, chiphi_output)
                        chiphi_wb = openpyxl.load_workbook(chiphi_output)
                        chiphi_ws = chiphi_wb.active
                        
                        # Áp dụng font 14 cho toàn bộ file chi phí
                        _apply_global_font(chiphi_ws, size=14)
                        
                        # Tự động tìm dòng trống đầu tiên trong file Chi phí (tránh ghi đè nếu template có data)
                        while chiphi_ws.cell(row=chiphi_current_row, column=1).value is not None or \
                              chiphi_ws.cell(row=chiphi_current_row, column=7).value is not None or \
                              chiphi_ws.cell(row=chiphi_current_row, column=8).value is not None:
                            chiphi_current_row += 1

                        # Khởi tạo Header cho cột Notes (Cột 15 - O)
                        cp_header_cell = chiphi_ws.cell(row=2, column=15, value="NOTES (Hệ thống OCR)")
                        # Header font size 14, bold, white
                        cp_header_cell.font = Font(color="FFFFFF", bold=True, size=14)
                        cp_header_cell.fill = _FILL_HEADER
                        cp_header_cell.alignment = Alignment(horizontal="center", vertical="center")
                        cp_header_cell.border = _THIN_BORDER
                        cp_notes_comment = Comment(_NOTES_CELL_COMMENT, "Lighthouse OCR")
                        cp_notes_comment.width = 450
                        cp_notes_comment.height = 200
                        cp_header_cell.comment = cp_notes_comment
                        chiphi_ws.column_dimensions['O'].width = 40

            if route_to_chiphi:
                # --- Ghi vào file Chi phí (15 cột + col 16 source filename) ---
                chiphi_row_data = {
                    1:  "NK",                       # A: MÃ CHỨNG TỪ
                    2:  _format_date(txn_info.get("invoice_date", "")), # B: NGÀY GHI SỔ (= ngày hoá đơn)
                    3:  _get_voucher(),               # C: C-3 lazy voucher
                    4:  bo_phan,                      # D: BỘ PHẬN
                    7:  supplier_code,                # G: MÃ ĐỐI TƯỢNG
                    8:  final_name,                   # H: DIỄN GIẢI (tên thô hoá đơn)
                    9:  6421,                         # I: TK NỢ
                    10: 331,                          # J: TK CÓ
                    11: "VND",                        # K: MÃ TIỀN TỆ
                    12: 1,                            # L: TỶ GIÁ
                    13: total_price_final,             # M: NGUYÊN TỆ
                    14: total_price_final,             # N: THÀNH TIỀN
                    15: notes_out,                     # O: NOTES
                    16: invoice_json.get("_source_filename", ""),  # P: source image for review panel
                }
                for col, value in chiphi_row_data.items():
                    cell = chiphi_ws.cell(row=chiphi_current_row, column=col, value=value)
                    if col in (13, 14) and isinstance(value, (int, float)):
                        cell.number_format = "#,##0"
                    cell.alignment = Alignment(vertical="center")
                    cell.border = _THIN_BORDER
                    cell.font = _FONT_DATA
                
                # Gắn comment hướng dẫn màu sắc vào header file Chi phí nếu là dòng đầu
                if chiphi_current_row == 3:
                    cp_comment = Comment(_NOTES_CELL_COMMENT, "Lighthouse OCR")
                    cp_comment.width = 450
                    cp_comment.height = 200
                    chiphi_ws.cell(row=1, column=8).comment = cp_comment

                # Màu sắc: Đỏ nếu high_risk, Vàng nếu bình thường (vì đã là hàng unmapped)
                cp_fill = _FILL_RED if high_risk else _FILL_YELLOW
                _apply_fill_to_row(chiphi_ws, chiphi_current_row, cp_fill, max_col=_CHIPHI_COLS)
                
                chiphi_current_row += 1
                chiphi_count += 1
            else:
                # --- Ghi vào PNMH (logic gốc 27 cột) ---
                row_data = {
                    1:  _format_date(txn_info.get("invoice_date", "")), # A: NGÀY GHI SỔ (= ngày hoá đơn)
                    2:  _get_voucher(),                                # B: C-3 lazy voucher
                    3:  _format_date(txn_info.get("invoice_date", "")), # C: NGÀY Hóa đơn
                    4:  None,                              # D: SỐ Hóa đơn
                    5:  None,                              # E: KÝ HIỆU
                    6:  bo_phan,                           # F: BỘ PHẬN
                    7:  None,                              # G: HỢP ĐỒNG
                    8:  supplier_code,                     # H: ĐỐI TƯỢNG (MÃ NCC)
                    9:  final_name,                        # I: DIỄN GIẢI
                    10: "331NK",                           # J: MÃ NHẬP XUẤT
                    11: data_store.kho_dict.get(          # K: KHO NHẬP (tra cứu theo bộ phận)
                            (bo_phan or "").strip().upper(),
                            "HH"                           # Fallback nếu không có trong Danh sách kho
                        ),
                    14: "VND",                             # N: MÃ TIỀN TỆ
                    15: 1,                                 # O: TỶ GIÁ
                    16: qty_final,                         # P: SỐ LƯỢNG
                    17: unit_price_final,                  # Q: ĐƠN GIÁ NGUYÊN TỆ
                    18: unit_price_final,                  # R: ĐƠN GIÁ THÀNH TIỀN
                    19: total_price_final,                 # S: THÀNH TIỀN NGUYÊN TỆ
                    20: total_price_final,                 # T: THÀNH TIỀN THÀNH TIỀN
                    27: notes_out,                         # AA: NOTES
                    28: invoice_json.get("_source_filename", ""),  # AB: Source image filename for preview
                }

                # M-4: Fallback rỗng thay vì hardcode mapped_code (tránh mã lỗi thời khi user sửa tên thủ công)
                vlookup_code = f"=IFERROR(VLOOKUP(I{current_row}, '{master_sheet_name}'!$A$1:$C${last_m_row}, 2, FALSE), \"\")"
                
                # Fix Unit-in-Name: Nếu unit_final (ví dụ dvt0) khác với dvt gốc của Master List, 
                # ta hardcode giá trị thay vì dùng VLOOKUP để tránh bị ghi đè lại thành dvt.
                master_dvt_check = m.get("unit", "") if m else ""
                if m and unit_final and master_dvt_check and unit_final.lower() != master_dvt_check.lower():
                    vlookup_unit = unit_final
                else:
                    vlookup_unit = f"=IFERROR(VLOOKUP(I{current_row}, '{master_sheet_name}'!$A$1:$C${last_m_row}, 3, FALSE), \"{unit_final}\")"
                
                row_data[12] = vlookup_code
                row_data[13] = vlookup_unit

                for col, value in row_data.items():
                    cell = ws.cell(row=current_row, column=col, value=value)
                    if col in (16, 17, 18, 19, 20) and isinstance(value, (int, float)):
                        cell.number_format = "#,##0.###" if col == 16 else "#,##0"
                    cell.alignment = Alignment(vertical="center")

                # --- Màu sắc theo ưu tiên ---
                # Đỏ: invoice high-risk
                # Vàng: item chưa khớp Master List (chỉ khi high_risk) HOẶC lỗi toán học
                # Xanh nhạt / Trắng: xen kẽ bình thường
                if high_risk:
                    row_fill = _FILL_RED
                elif math_error or item_unmapped or "[MIXED_UNIT]" in notes_out:
                    row_fill = _FILL_YELLOW
                else:
                    row_fill = base_fill
                _apply_fill_to_row(ws, current_row, row_fill)

                current_row += 1

    # Save Chi phí file nếu có dữ liệu
    if chiphi_wb is not None and chiphi_count > 0:
        try:
            _apply_global_font(chiphi_ws, size=14)
            chiphi_ws.freeze_panes = 'A3'
            _auto_fit_columns(chiphi_ws)
            chiphi_wb.save(chiphi_output)
            if status_callback:
                status_callback(f"📊 Đã ghi {chiphi_count} item chi phí vào: {os.path.basename(chiphi_output)}")
        except PermissionError:
            if status_callback:
                status_callback(f"⚠️ Không thể lưu file Chi phí — file đang mở?")
    elif chiphi_wb is not None:
        # Template đã copy nhưng không có data → xóa file rỗng
        try:
            os.remove(chiphi_output)
        except OSError:
            pass

    try:
        # H-6: Chỉ thực hiện định dạng và AutoFit SAU KHI đã paste xong toàn bộ dữ liệu
        _apply_global_font(ws, size=14)
        ws.freeze_panes = 'A3'
        _auto_fit_columns(ws)
        wb.save(output_file)
    except PermissionError:
        raise PermissionError(
            f"⚠️ Không thể lưu file Excel vào thư mục DONE!\n\n"
            f"Một file tên '{os.path.basename(output_file)}' đang bị mở trong Excel."
        )

    if chiphi_wb is not None and chiphi_count > 0:
        return f"{output_file}|{chiphi_output}"
    return output_file
