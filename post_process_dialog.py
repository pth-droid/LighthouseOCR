import os
import sys
import shutil
import copy
import json
from PIL import Image as PilImage
import openpyxl
import datetime
import re
from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QLabel, QTabWidget, QWidget, QCheckBox, QMessageBox,
    QHeaderView, QAbstractItemView, QStyledItemDelegate, QComboBox, QApplication, QCompleter,
    QSplitter, QScrollArea, QFrame, QSizePolicy
)
from PyQt5.QtCore import Qt, QEvent, QPoint, QSettings, QTimer, pyqtSignal
from PyQt5.QtGui import QColor, QKeyEvent, QTextCursor, QPixmap
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from path_utils import get_root_dir, get_asset_path

# ──────────────────────────────────────────────
#  Constants
# ──────────────────────────────────────────────
_ROOT_DIR              = get_root_dir()
_CHIPHI_TEMPLATE_FILE  = os.path.join(_ROOT_DIR, "Data structure", "Mẫu Nhập Chi phí.xlsx")

_DEPT_OPTIONS = ["BAR", "BEP", "BANH", "RANG"]

# PNMH: Các cột hiển thị trong dialog (column index PNMH 1-based + label)
_PNMH_COLS_INFO = [
    (2, "SỐ CHỨNG TỪ"),
    (3, "NGÀY HOÁ ĐƠN"),
    (6, "BỘ PHẬN"),
    (8, "ĐỐI TƯỢNG"),
    (9, "DIỄN GIẢI"),
    (11, "KHO NHẬP"),
    (12, "MÃ VẬT TƯ"),
    (13, "ĐƠN VỊ TÍNH"),
    (16, "SỐ LƯỢNG"),
    (17, "ĐƠN GIA"),
    (20, "THÀNH TIỀN"),
    (27, "GHI CHÚ")
]

# Chi phí: Các cột hiển thị trong dialog
_CHIPHI_COLS_INFO = [
    (2, "NGÀY GHI SỔ"),
    (3, "SỐ CHỨNG TỪ"),
    (4, "BỘ PHẬN"),
    (7, "MÃ ĐỐI TƯỢNG"),
    (8, "DIỄN GIẢI"),
    (9, "TK NỢ"),
    (10, "TK CÓ"),
    (14, "THÀNH TIỀN"),
    (15, "GHI CHÚ")
]

# Màu sắc UI
_COLOR_TEXT     = QColor("#BDD8E9")

# Style cho openpyxl
_SIDE = Side(style="thin", color="000000")
_THIN_BORDER = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)
_FONT_DATA = Font(name="Calibri", size=14)

def _fmt_money(val):
    if val is None: return ""
    try:
        return "{:,.0f}".format(float(val))
    except:
        return str(val)

def _fmt_date(val):
    if not val: return ""
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val.strftime("%d/%m/%Y")
    
    sval = str(val).strip()
    if not sval: return ""
    
    # Hỗ trợ gõ tắt: ddmmyy, ddmmyyyy, ddmm
    now = datetime.datetime.now()
    year = now.year
    
    clean_val = sval.replace("/", "").replace("-", "").replace(".", "")
    if clean_val.isdigit():
        if len(clean_val) == 4: # ddmm -> dd/mm/yyyy
            sval = f"{clean_val[:2]}/{clean_val[2:]}/{year}"
        elif len(clean_val) == 6: # ddmmyy -> dd/mm/20yy
            sval = f"{clean_val[:2]}/{clean_val[2:4]}/20{clean_val[4:]}"
        elif len(clean_val) == 8: # ddmmyyyy -> dd/mm/yyyy
            sval = f"{clean_val[:2]}/{clean_val[2:4]}/{clean_val[4:]}"

    if " " in sval:
        sval = sval.split(" ")[0]

    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.datetime.strptime(sval, fmt).strftime("%d/%m/%Y")
        except:
            continue
    return sval

class PasteableTableWidget(QTableWidget):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._dragging = False
        self._last_pos = None

    def mousePressEvent(self, event):
        if event.button() == Qt.MiddleButton or (event.button() == Qt.LeftButton and not self.itemAt(event.pos())):
            self._dragging = True
            self._last_pos = event.pos()
            self.setCursor(Qt.ClosedHandCursor)
            event.accept()
        else:
            super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self._dragging and self._last_pos:
            delta = event.pos() - self._last_pos
            # Giảm tốc độ grip chuột kéo cuộn (0.6x)
            self.horizontalScrollBar().setValue(self.horizontalScrollBar().value() - int(delta.x() * 0.6))
            self.verticalScrollBar().setValue(self.verticalScrollBar().value() - int(delta.y() * 0.6))
            self._last_pos = event.pos()
            event.accept()
        else:
            super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        if self._dragging:
            self._dragging = False
            self.setCursor(Qt.ArrowCursor)
            event.accept()
        else:
            super().mouseReleaseEvent(event)

    def keyPressEvent(self, event: QKeyEvent):
        if event.modifiers() & Qt.ControlModifier and event.key() == Qt.Key_V:
            self._paste_data()
        else:
            super().keyPressEvent(event)

    def _paste_data(self):
        clipboard = QApplication.clipboard()
        text = clipboard.text()
        if not text:
            return
            
        selected_ranges = self.selectedRanges()
        if not selected_ranges:
            return

        # Tách dữ liệu clipboard thành mảng 2D
        rows_data = [r.split('\t') for r in text.strip().split('\n')]
        row_count_data = len(rows_data)
        col_count_data = len(rows_data[0]) if row_count_data > 0 else 0
        
        # Lấy danh sách tất cả các ô được chọn (row, col)
        selected_items = []
        for rng in selected_ranges:
            for r in range(rng.topRow(), rng.bottomRow() + 1):
                for c in range(rng.leftColumn(), rng.rightColumn() + 1):
                    # Bỏ qua cột checkbox (col 0)
                    if c > 0:
                        selected_items.append((r, c))
        
        if not selected_items:
            return

        # Case 1: Clipboard chỉ có 1 ô đơn lẻ -> Paste vào tất cả các ô đang chọn
        if row_count_data == 1 and col_count_data == 1:
            val = rows_data[0][0].strip()
            for r, c in selected_items:
                self._set_cell_text(r, c, val)
        else:
            # Case 2: Clipboard có nhiều ô -> Paste theo offset từ ô đầu tiên được chọn
            base_r, base_c = selected_items[0]
            for r_idx, row_vals in enumerate(rows_data):
                for c_idx, col_text in enumerate(row_vals):
                    target_r = base_r + r_idx
                    target_c = base_c + c_idx
                    if target_r < self.rowCount() and target_c < self.columnCount() and target_c > 0:
                        self._set_cell_text(target_r, target_c, col_text.strip())

    def _set_cell_text(self, r, c, text):
        self.blockSignals(True)
        item = self.item(r, c)
        if item:
            item.setText(text)
        else:
            new_item = QTableWidgetItem(text)
            new_item.setForeground(_COLOR_TEXT)
            self.setItem(r, c, new_item)
        self.blockSignals(False)

# ── Widget bao quanh Checkbox để tăng diện tích click ──
class ClickableCheckBoxWidget(QWidget):
    def __init__(self, checkbox, bg_color=None, parent=None):
        super().__init__(parent)
        self.checkbox = checkbox
        self.setAutoFillBackground(True)
        if bg_color:
            p = self.palette()
            p.setColor(self.backgroundRole(), bg_color)
            self.setPalette(p)
            
        layout = QHBoxLayout(self)
        layout.addWidget(self.checkbox)
        layout.setAlignment(Qt.AlignCenter)
        layout.setContentsMargins(0, 0, 0, 0)
        self.setLayout(layout)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.checkbox.toggle()
        super().mousePressEvent(event)

# ── Delegate cho Combobox có thể gõ ──
class ComboBoxDelegate(QStyledItemDelegate):
    def __init__(self, items, parent=None):
        super().__init__(parent)
        self.items = items

    def createEditor(self, parent, option, index):
        editor = QComboBox(parent)
        items_list = self.items(index) if callable(self.items) else self.items
        editor.addItems(items_list)
        editor.setEditable(True)
        # Tắt tự động điền để không bị nhảy chữ khi gõ
        editor.setInsertPolicy(QComboBox.NoInsert)
        
        # Thiết lập QCompleter để hỗ trợ gợi ý khi gõ
        completer = QCompleter(items_list, editor)
        completer.setCaseSensitivity(Qt.CaseInsensitive)
        completer.setFilterMode(Qt.MatchContains) # Tìm kiếm chuỗi con
        completer.setCompletionMode(QCompleter.PopupCompletion)
        editor.setCompleter(completer)
        
        editor.setStyleSheet("""
            QComboBox { background:#0A2740; color:#BDD8E9; border:1px solid #49769F; padding:2px; }
            QComboBox QAbstractItemView { background:#0A2740; color:#BDD8E9; selection-background-color:#0A4174; }
        """)
        
        return editor

    def setEditorData(self, editor, index):
        value = index.model().data(index, Qt.EditRole)
        text = str(value) if value is not None else ""
        editor.setCurrentText(text)
        
        # Tự động chọn toàn bộ text.
        # QCompleter sẽ tự động hiện popup danh sách gợi ý khi người dùng bắt đầu gõ.
        line_edit = editor.lineEdit()
        if line_edit:
            line_edit.selectAll()

    def setModelData(self, editor, model, index):
        value = editor.currentText()
        model.setData(index, value, Qt.EditRole)


# ── Helpers ──────────────────────────────────────────────────────────────────
def _append_to_chiphi(chiphi_ws, chiphi_row: int, pnmh_row_data: dict, kho_dict: dict,
                      source_filename: str = "", invoice_date: str = ""):
    date_val  = invoice_date or ""
    bo_phan   = str(pnmh_row_data.get(6, "") or "").strip()

    chiphi_data = {
        1:  "NK",
        2:  date_val,
        3:  pnmh_row_data.get(2, ""),
        4:  bo_phan,
        7:  pnmh_row_data.get(8, ""),
        8:  pnmh_row_data.get(9, ""),
        9:  6421,
        10: 331,
        11: "VND",
        12: 1,
        13: pnmh_row_data.get(20, None),
        14: pnmh_row_data.get(20, None),
        15: f"[Chuyển từ PNMH] {pnmh_row_data.get(27, '') or ''}".strip(),
        16: source_filename,  # Source image filename for review panel
    }

    for col, value in chiphi_data.items():
        cell = chiphi_ws.cell(row=chiphi_row, column=col, value=value)
        if col in (13, 14) and isinstance(value, (int, float)):
            cell.number_format = "#,##0"
        cell.alignment = Alignment(vertical="center")
        cell.border = _THIN_BORDER
        cell.font = _FONT_DATA
        
    chiphi_ws.row_dimensions[chiphi_row].height = 18

def _find_first_empty_row(ws, start_row: int = 3) -> int:
    row = start_row
    while True:
        if all(ws.cell(row=row, column=c).value is None for c in [1, 7, 8]):
            return row
        row += 1


# ════════════════════════════════════════════════════════════════════════════
class PostProcessDialog(QDialog):
    def __init__(self, pnmh_path: str, chiphi_path: str | None,
                 app_data, parent=None):
        super().__init__(parent)
        self.pnmh_path   = pnmh_path
        self.chiphi_path = chiphi_path
        self.app_data    = app_data
        self.kho_dict    = app_data.kho_dict

        # Dữ liệu PNMH
        self._pnmh_rows: list[dict] = []
        self._pnmh_row_indices: list[int] = []
        self._deleted_pnmh_ws_rows: list[int] = []
        self._pnmh_image_filenames: list[str] = []  # col 28 from PNMH Excel

        # Dữ liệu Chi phí
        self._chiphi_rows: list[dict] = []
        self._chiphi_row_indices: list[int] = []
        self._deleted_chiphi_ws_rows: list[int] = []
        self._chiphi_image_filenames: list[str] = []  # col 16 from ChiPhi Excel
        self._pnmh_original_by_ws_row: dict[int, dict] = {}
        self._chiphi_original_by_ws_row: dict[int, dict] = {}

        # Image viewer state per panel (keyed by 'pnmh' and 'chiphi')
        self._img_state: dict = {
            'pnmh':   {'zoom': 1.0, 'path': '', 'pan': False, 'pan_pos': QPoint()},
            'chiphi': {'zoom': 1.0, 'path': '', 'pan': False, 'pan_pos': QPoint()},
        }

        self.setWindowTitle("Rà soát File Kết Quả")
        self.setStyleSheet(self._stylesheet())
        self.finished.connect(self._save_settings)

        if not self.chiphi_path:
            pnmh_dir  = os.path.dirname(self.pnmh_path)
            pnmh_name = os.path.splitext(os.path.basename(self.pnmh_path))[0]
            chiphi_name = pnmh_name.replace("PNMH", "ChiPhi") + "_ChiPhi.xlsx"
            self.chiphi_path = os.path.join(pnmh_dir, chiphi_name)

        self._build_ui()
        self._load_pnmh()
        self._load_chiphi()
        self._load_settings()  # Restore geometry/splitter after UI is built

    def keyPressEvent(self, event):
        # Bỏ tính năng ấn Esc đóng bảng review
        if event.key() == Qt.Key_Escape:
            event.ignore()
        else:
            super().keyPressEvent(event)

    def showEvent(self, event):
        super().showEvent(event)
        QTimer.singleShot(0, self._apply_initial_splitter_sizes)

    def _apply_initial_splitter_sizes(self):
        settings = QSettings("LighthouseOCR", "PostProcessDialog")
        for attr, key, default_ratio in [
            ('pnmh_splitter',   'pnmh_splitter',   (0.65, 0.35)),
            ('chiphi_splitter', 'chiphi_splitter',  (0.65, 0.35)),
        ]:
            splitter = getattr(self, attr, None)
            if splitter is None:
                continue
            saved = settings.value(key)
            if saved:
                try:
                    splitter.setSizes([int(s) for s in saved])
                    continue
                except (ValueError, TypeError):
                    pass
            total = splitter.width()
            if total > 0:
                splitter.setSizes([int(total * default_ratio[0]), int(total * default_ratio[1])])

    def _load_settings(self):
        settings = QSettings("LighthouseOCR", "PostProcessDialog")
        maximized = settings.value("maximized", True, type=bool)
        if maximized:
            self.setWindowState(Qt.WindowMaximized)
        else:
            geom = settings.value("geometry")
            if geom:
                self.restoreGeometry(geom)
            else:
                self.setWindowState(Qt.WindowMaximized)

    def _save_settings(self):
        settings = QSettings("LighthouseOCR", "PostProcessDialog")
        is_max = bool(self.windowState() & Qt.WindowMaximized)
        settings.setValue("maximized", is_max)
        if not is_max:
            settings.setValue("geometry", self.saveGeometry())
        for attr, key in [('pnmh_splitter', 'pnmh_splitter'), ('chiphi_splitter', 'chiphi_splitter')]:
            splitter = getattr(self, attr, None)
            if splitter:
                settings.setValue(key, splitter.sizes())

    # ── UI ──────────────────────────────────────────────────────────────────
    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setSpacing(10)
        root.setContentsMargins(14, 14, 14, 14)

        self.tabs = QTabWidget()
        self.tab_pnmh = QWidget()
        self.tab_chiphi = QWidget()
        self.tabs.addTab(self.tab_pnmh, "📁 Phiếu Nhập Mua Hàng")
        self.tabs.addTab(self.tab_chiphi, "📁 Nhập Chi Phí")
        
        self._build_pnmh_tab()
        self._build_chiphi_tab()
        
        root.addWidget(self.tabs, stretch=1)

        # Status bar: left=column name, right=full selected cell content
        pps_status_w = QWidget()
        pps_status_w.setFixedHeight(26)
        pps_status_w.setStyleSheet(
            "background:rgba(10,39,64,0.5); border-radius:3px;"
            " border:1px solid rgba(123,189,232,0.12);"
        )
        pps_sl = QHBoxLayout(pps_status_w)
        pps_sl.setContentsMargins(8, 0, 8, 0)
        pps_sl.setSpacing(10)
        self.lbl_pps_col = QLabel("")
        self.lbl_pps_col.setStyleSheet(
            "color:#7BBDE8; font-size:12px; border:none; background:transparent;"
        )
        self.lbl_pps_cell = QLabel("")
        self.lbl_pps_cell.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.lbl_pps_cell.setStyleSheet(
            "color:#BDD8E9; font-size:12px; border:none; background:transparent;"
        )
        pps_sl.addWidget(self.lbl_pps_col, 1)
        pps_sl.addWidget(self.lbl_pps_cell, 2)
        root.addWidget(pps_status_w)

        # Bottom Actions
        footer = QHBoxLayout()
        self.btn_delete = QPushButton("🗑️  Xoá")
        self.btn_delete.setObjectName("btn_delete")
        self.btn_delete.setFixedWidth(100)
        self.btn_delete.clicked.connect(self._do_delete)

        self.btn_save = QPushButton("💾  Lưu & Đóng")
        self.btn_save.setObjectName("btn_save")
        self.btn_save.setFixedWidth(180)
        self.btn_save.clicked.connect(self._do_save_all_and_close)
        
        self.btn_cancel = QPushButton("Đóng (Không lưu)")
        self.btn_cancel.setFixedWidth(160)
        self.btn_cancel.clicked.connect(self.reject)

        footer.addWidget(self.btn_delete)
        footer.addStretch()
        footer.addWidget(self.btn_cancel)
        footer.addWidget(self.btn_save)
        root.addLayout(footer)

        # Wire both tables to update status bar on cell selection
        self.pnmh_table.currentCellChanged.connect(
            lambda r, c, _pr, _pc: self._update_status_bar(self.pnmh_table, r, c)
        )
        self.chiphi_table.currentCellChanged.connect(
            lambda r, c, _pr, _pc: self._update_status_bar(self.chiphi_table, r, c)
        )

    def _build_image_panel(self, scroll_attr: str, label_attr: str) -> QWidget:
        """Factory: builds a right-side invoice image panel. Stores scroll/label as self.<attr>."""
        # Derive panel_id: 'pnmh' or 'chiphi' from scroll_attr
        panel_id = 'chiphi' if 'chiphi' in scroll_attr else 'pnmh'

        panel = QWidget()
        panel.setStyleSheet(
            "background: rgba(0,10,20,0.6); border: 1px solid rgba(123,189,232,0.2); border-radius: 4px;"
        )
        pl = QVBoxLayout(panel)
        pl.setContentsMargins(6, 6, 6, 6)
        pl.setSpacing(4)

        # ── Top bar: title + rotate buttons ──
        top_bar = QHBoxLayout()
        top_bar.setSpacing(6)

        title = QLabel("🖼  Ảnh Hóa Đơn  |  Cuộn: zoom  |  Kéo: di chuyển")
        title.setAlignment(Qt.AlignVCenter | Qt.AlignLeft)
        title.setStyleSheet("color: #7BBDE8; font-weight: 700; font-size: 11px; border: none; background: transparent;")
        top_bar.addWidget(title, stretch=1)

        _btn_style = (
            "QPushButton { background: rgba(10,65,116,0.5); color: #7BBDE8; border: 1px solid rgba(123,189,232,0.25);"
            " border-radius: 5px; padding: 2px 8px; font-size: 14px; min-height: 22px; min-width: 28px; }"
            "QPushButton:hover { background: rgba(73,118,159,0.6); color: #ffffff; }"
            "QPushButton:pressed { background: rgba(10,65,116,0.9); }"
        )
        btn_left = QPushButton("↺")
        btn_left.setToolTip("Xoay trái 90°")
        btn_left.setStyleSheet(_btn_style)
        btn_left.setFixedSize(30, 24)
        btn_left.clicked.connect(lambda: self._rotate_invoice_image(panel_id, 90))
        top_bar.addWidget(btn_left)

        btn_right = QPushButton("↻")
        btn_right.setToolTip("Xoay phải 90°")
        btn_right.setStyleSheet(_btn_style)
        btn_right.setFixedSize(30, 24)
        btn_right.clicked.connect(lambda: self._rotate_invoice_image(panel_id, -90))
        top_bar.addWidget(btn_right)

        pl.addLayout(top_bar)

        scroll = QScrollArea()
        scroll.setWidgetResizable(False)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }")
        scroll.setMinimumWidth(200)

        label = QLabel("Chọn một dòng\nđể xem ảnh hóa đơn")
        label.setAlignment(Qt.AlignTop | Qt.AlignHCenter)
        label.setWordWrap(True)
        label.setStyleSheet("color: #49769F; font-size: 12px; background: transparent;")
        scroll.setWidget(label)
        scroll.viewport().installEventFilter(self)

        pl.addWidget(scroll, stretch=1)

        setattr(self, scroll_attr, scroll)
        setattr(self, label_attr, label)
        return panel


    def _build_pnmh_tab(self):
        layout = QVBoxLayout(self.tab_pnmh)

        header = QHBoxLayout()
        header.addWidget(QLabel("<b>Dữ liệu Phiếu Nhập Mua Hàng (PNMH)</b>"))
        header.addStretch()

        self.btn_transfer = QPushButton("➡  Chuyển sang Chi phí")
        self.btn_transfer.setObjectName("btn_transfer")
        self.btn_transfer.setEnabled(False)
        self.btn_transfer.clicked.connect(self._do_transfer)
        header.addWidget(self.btn_transfer)

        self.btn_add_pnmh = QPushButton("+ Thêm Dòng")
        self.btn_add_pnmh.setFixedWidth(120)
        self.btn_add_pnmh.clicked.connect(self._do_add_pnmh_row)
        header.addWidget(self.btn_add_pnmh)
        layout.addLayout(header)

        # ── Horizontal splitter: table (left) + image panel (right) ──
        self.pnmh_splitter = QSplitter(Qt.Horizontal)
        self.pnmh_splitter.setHandleWidth(6)
        self.pnmh_splitter.setStyleSheet(
            "QSplitter::handle { background: rgba(123,189,232,0.15); border-radius: 3px; }"
        )

        # ── Left: data table ──
        self.pnmh_table = PasteableTableWidget()
        self.pnmh_table.setColumnCount(len(_PNMH_COLS_INFO) + 1)
        col_labels = ["X"] + [lbl for _, lbl in _PNMH_COLS_INFO]
        self.pnmh_table.setHorizontalHeaderLabels(col_labels)

        self.pnmh_table.setEditTriggers(
            QAbstractItemView.AnyKeyPressed |
            QAbstractItemView.DoubleClicked |
            QAbstractItemView.SelectedClicked
        )

        # Nạp dữ liệu danh mục cho Dropdowns
        supplier_list = []
        if self.app_data:
            for code, name in self.app_data.suppliers_dict.items():
                supplier_list.extend([str(code), str(name)])
        supplier_list = sorted(list(set(supplier_list)))
        item_list     = sorted([str(v.get("name", "")) for v in self.app_data.items_dict.values()]) if self.app_data else []
        kho_list      = sorted(list(set([str(v) for v in self.app_data.kho_dict.values()]))) if self.app_data else []

        # Combo Delegate cho các cột PNMH (Index UI 1-based, bỏ qua chk ở 0)
        # 3: BỘ PHẬN | 4: ĐỐI TƯỢNG | 5: DIỄN GIẢI | 6: KHO NHẬP | 8: ĐVT
        self.pnmh_table.setItemDelegateForColumn(3, ComboBoxDelegate(_DEPT_OPTIONS, self.pnmh_table))
        self.pnmh_table.setItemDelegateForColumn(4, ComboBoxDelegate(supplier_list, self.pnmh_table))
        # Cột Diễn giải (5) dùng item_list chung
        self.pnmh_table.setItemDelegateForColumn(5, ComboBoxDelegate(item_list, self.pnmh_table))
        self.pnmh_table.setItemDelegateForColumn(6, ComboBoxDelegate(kho_list, self.pnmh_table))
        # Cột ĐVT (8) dùng dropdown động theo tên hàng
        self.pnmh_table.setItemDelegateForColumn(8, ComboBoxDelegate(self._get_dvt_options, self.pnmh_table))

        self.pnmh_table.cellChanged.connect(self._on_pnmh_cell_changed)

        # Bật menu chuột phải cho PNMH
        self.pnmh_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.pnmh_table.customContextMenuRequested.connect(self._show_pnmh_context_menu)

        # Selection signal: hiển thị ảnh hóa đơn tương ứng
        self.pnmh_table.selectionModel().selectionChanged.connect(self._on_pnmh_selection_changed)

        self._setup_table_style(self.pnmh_table)
        self.pnmh_splitter.addWidget(self.pnmh_table)

        # ── Right: invoice image panel ──
        self.pnmh_splitter.addWidget(self._build_image_panel('img_scroll', 'img_label'))

        # ── Resize weights: table 65%, image panel 35% ──
        self.pnmh_splitter.setStretchFactor(0, 65)
        self.pnmh_splitter.setStretchFactor(1, 35)

        layout.addWidget(self.pnmh_splitter, stretch=1)

    def _get_dvt_options(self, index):
        row = index.row()
        item = self.pnmh_table.item(row, 5) # 5 là cột Diễn giải
        if item and self.app_data:
            diengiai = item.text().lower().strip()
            all_units = set()
            
            # 1. Lấy đơn vị từ Master List
            if hasattr(self.app_data, 'items_dict') and diengiai in self.app_data.items_dict:
                info = self.app_data.items_dict[diengiai]
                if info.get("unit"): all_units.add(info["unit"])
                if info.get("dvt0"): all_units.add(info["dvt0"])
            
            # 2. Lấy đơn vị từ Alias Dictionary (Cấu trúc 1-N mới)
            if hasattr(self.app_data, 'aliases_dict') and diengiai in self.app_data.aliases_dict:
                alias_info = self.app_data.aliases_dict[diengiai]
                for u_info in alias_info.get("units", []):
                    if u_info.get("unit"):
                        all_units.add(u_info["unit"])
            
            if all_units:
                return sorted(list(all_units))
        return []

    def _on_pnmh_selection_changed(self):
        indexes = self.pnmh_table.selectedIndexes()
        if not indexes:
            return
        row = min(idx.row() for idx in indexes)
        if row >= len(self._pnmh_image_filenames) or not self._pnmh_image_filenames[row]:
            self._clear_image('pnmh', "Không có ảnh\ncho dòng này")
            return
        self._img_state['pnmh']['zoom'] = 1.0
        img_path = os.path.join(os.path.dirname(self.pnmh_path), self._pnmh_image_filenames[row])
        self._display_invoice_image('pnmh', img_path)

    def _on_chiphi_selection_changed(self):
        indexes = self.chiphi_table.selectedIndexes()
        if not indexes:
            return
        row = min(idx.row() for idx in indexes)
        if row >= len(self._chiphi_image_filenames) or not self._chiphi_image_filenames[row]:
            self._clear_image('chiphi', "Không có ảnh\ncho dòng này")
            return
        self._img_state['chiphi']['zoom'] = 1.0
        img_path = os.path.join(os.path.dirname(self.pnmh_path), self._chiphi_image_filenames[row])
        self._display_invoice_image('chiphi', img_path)

    def _clear_image(self, panel_id: str, message: str):
        scroll, label = self._get_img_widgets(panel_id)
        label.setPixmap(QPixmap())
        label.setText(message)
        label.resize(scroll.viewport().size())
        self._img_state[panel_id]['path'] = ''

    def _rotate_invoice_image(self, panel_id: str, degrees: int):
        """Rotate the on-disk image by degrees (positive = CCW) and refresh the panel."""
        state = self._img_state[panel_id]
        path = state.get('path', '')
        if not path or not os.path.exists(path):
            return
        try:
            img = PilImage.open(path)
            rotated = img.rotate(degrees, expand=True)
            img.close()
            ext = os.path.splitext(path)[1].lower()
            if ext in ('.jpg', '.jpeg'):
                rotated.save(path, quality=95)
            else:
                rotated.save(path)
            state['zoom'] = 1.0
            self._display_invoice_image(panel_id, path)
        except Exception as e:
            from PyQt5.QtWidgets import QMessageBox
            QMessageBox.warning(self, "Lỗi xoay ảnh", str(e))

    def _display_invoice_image(self, panel_id: str, img_path: str):
        scroll, label = self._get_img_widgets(panel_id)
        state = self._img_state[panel_id]
        try:
            if not os.path.exists(img_path):
                self._clear_image(panel_id, f"Không tìm thấy ảnh:\n{img_path}")
                return
            pix = QPixmap(img_path)
            if pix.isNull():
                self._clear_image(panel_id, "Không thể tải ảnh")
                return
            state['path'] = img_path
            base_w = max(scroll.viewport().width() - 4, 100)
            target_w = max(int(base_w * state['zoom']), 50)
            scaled = pix.scaledToWidth(target_w, Qt.SmoothTransformation)
            label.setText("")
            label.setPixmap(scaled)
            label.resize(scaled.size())
        except Exception:
            self._clear_image(panel_id, "Lỗi hiển thị ảnh")

    def _panel_id_for_viewport(self, obj) -> str | None:
        if hasattr(self, 'img_scroll') and obj is self.img_scroll.viewport():
            return 'pnmh'
        if hasattr(self, 'chiphi_img_scroll') and obj is self.chiphi_img_scroll.viewport():
            return 'chiphi'
        return None

    def _get_img_widgets(self, panel_id: str):
        if panel_id == 'pnmh':
            return self.img_scroll, self.img_label
        return self.chiphi_img_scroll, self.chiphi_img_label

    def eventFilter(self, obj, event):
        pid = self._panel_id_for_viewport(obj)
        if pid is None:
            return super().eventFilter(obj, event)

        scroll, label = self._get_img_widgets(pid)
        state = self._img_state[pid]
        etype = event.type()

        if etype == QEvent.Wheel:
            delta = event.angleDelta().y()
            factor = 1.15 if delta > 0 else (1.0 / 1.15)
            state['zoom'] = max(0.2, min(8.0, state['zoom'] * factor))
            if state['path']:
                self._display_invoice_image(pid, state['path'])
            return True  # Consume; don't propagate to splitter

        if etype == QEvent.MouseButtonPress and event.button() == Qt.LeftButton:
            state['pan'] = True
            state['pan_pos'] = event.pos()
            scroll.viewport().setCursor(Qt.ClosedHandCursor)
            return True

        if etype == QEvent.MouseMove and state['pan']:
            delta = event.pos() - state['pan_pos']
            state['pan_pos'] = event.pos()
            scroll.horizontalScrollBar().setValue(
                scroll.horizontalScrollBar().value() - delta.x()
            )
            scroll.verticalScrollBar().setValue(
                scroll.verticalScrollBar().value() - delta.y()
            )
            return True

        if etype == QEvent.MouseButtonRelease and event.button() == Qt.LeftButton:
            state['pan'] = False
            scroll.viewport().setCursor(Qt.ArrowCursor)
            return True

        if etype == QEvent.ContextMenu:
            if state['path'] and os.path.exists(state['path']):
                import subprocess
                from PyQt5.QtWidgets import QMenu, QAction
                menu = QMenu(self)
                act = QAction("📂  Mở thư mục chứa ảnh", self)
                norm_path = os.path.normpath(os.path.abspath(state['path']))
                act.triggered.connect(lambda _=False, p=norm_path: subprocess.run(
                    ['explorer', '/select,', p], check=False
                ))
                menu.addAction(act)
                menu.addSeparator()
                act_left = QAction("↩  Xoay trái 90°", self)
                act_left.triggered.connect(lambda _=False, p=pid: self._rotate_invoice_image(p, 90))
                menu.addAction(act_left)
                act_right = QAction("↪  Xoay phải 90°", self)
                act_right.triggered.connect(lambda _=False, p=pid: self._rotate_invoice_image(p, -90))
                menu.addAction(act_right)
                menu.exec_(event.globalPos())
            return True

        return super().eventFilter(obj, event)

    def _build_chiphi_tab(self):
        layout = QVBoxLayout(self.tab_chiphi)
        header = QHBoxLayout()
        header.addWidget(QLabel("<b>Dữ liệu Nhập Chi Phí (Dùng cho hạch toán Nợ 6421 / Có 331)</b>"))
        header.addStretch()
        self.btn_add_chiphi = QPushButton("+ Thêm Dòng")
        self.btn_add_chiphi.setFixedWidth(120)
        self.btn_add_chiphi.clicked.connect(self._do_add_chiphi_row)
        header.addWidget(self.btn_add_chiphi)
        layout.addLayout(header)

        # ── Horizontal splitter: table (left) + image panel (right) ──
        self.chiphi_splitter = QSplitter(Qt.Horizontal)
        self.chiphi_splitter.setHandleWidth(6)
        self.chiphi_splitter.setStyleSheet(
            "QSplitter::handle { background: rgba(123,189,232,0.15); border-radius: 3px; }"
        )

        self.chiphi_table = PasteableTableWidget()
        self.chiphi_table.setColumnCount(len(_CHIPHI_COLS_INFO))
        col_labels = [lbl for _, lbl in _CHIPHI_COLS_INFO]
        self.chiphi_table.setHorizontalHeaderLabels(col_labels)

        self.chiphi_table.setEditTriggers(
            QAbstractItemView.AnyKeyPressed |
            QAbstractItemView.DoubleClicked |
            QAbstractItemView.SelectedClicked
        )

        # Nạp dữ liệu danh mục cho Dropdowns (Chi phí)
        supplier_list = []
        if self.app_data:
            for code, name in self.app_data.suppliers_dict.items():
                supplier_list.extend([str(code), str(name)])
        supplier_list = sorted(list(set(supplier_list)))
        item_list     = sorted([str(v.get("name", "")) for v in self.app_data.items_dict.values()]) if self.app_data else []

        # Combo Delegate cho các cột Chi phí (Index UI 0-based)
        # 2: BỘ PHẬN | 3: MÃ ĐỐI TƯỢNG | 4: DIỄN GIẢI
        self.chiphi_table.setItemDelegateForColumn(2, ComboBoxDelegate(_DEPT_OPTIONS, self.chiphi_table))
        self.chiphi_table.setItemDelegateForColumn(3, ComboBoxDelegate(supplier_list, self.chiphi_table))
        self.chiphi_table.setItemDelegateForColumn(4, ComboBoxDelegate(item_list, self.chiphi_table))

        self.chiphi_table.cellChanged.connect(self._on_chiphi_cell_changed)
        self.chiphi_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.chiphi_table.customContextMenuRequested.connect(self._show_chiphi_context_menu)
        self.chiphi_table.selectionModel().selectionChanged.connect(self._on_chiphi_selection_changed)

        self._setup_table_style(self.chiphi_table)
        self.chiphi_splitter.addWidget(self.chiphi_table)

        # ── Right: invoice image panel ──
        self.chiphi_splitter.addWidget(self._build_image_panel('chiphi_img_scroll', 'chiphi_img_label'))

        self.chiphi_splitter.setStretchFactor(0, 65)
        self.chiphi_splitter.setStretchFactor(1, 35)

        layout.addWidget(self.chiphi_splitter, stretch=1)

    def _update_status_bar(self, table: QTableWidget, row: int, col: int):
        if row < 0 or col < 0:
            return
        item = table.item(row, col)
        text = item.text() if item else ""
        header = table.horizontalHeaderItem(col)
        col_name = header.text() if header else ""
        self.lbl_pps_col.setText(col_name)
        self.lbl_pps_cell.setText(text)

    def _setup_table_style(self, table: QTableWidget):
        table.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        table.horizontalHeader().setStretchLastSection(True) # Đảm bảo bảng phủ hết chiều rộng
        table.setSelectionBehavior(QAbstractItemView.SelectItems)
        table.setAlternatingRowColors(False)
        table.verticalHeader().setVisible(False)
        table.setShowGrid(True)
        
    # ── Load data ──
    def _load_pnmh(self):
        self.pnmh_table.setRowCount(0)
        self._pnmh_rows.clear()
        self._pnmh_row_indices.clear()
        self._pnmh_image_filenames.clear()

        if not os.path.exists(self.pnmh_path): return

        try:
            wb = openpyxl.load_workbook(self.pnmh_path, data_only=True)
            ws = wb.active
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
                if not any(row): continue
                
                row_vals = {c_idx+1: val for c_idx, val in enumerate(row)}
                
                # Logic vlookup bằng Python
                if not row_vals.get(12) or not row_vals.get(13):
                    diengiai = row_vals.get(9, "")
                    if diengiai and self.app_data and hasattr(self.app_data, 'items_dict'):
                        item_info = self.app_data.items_dict.get(str(diengiai).lower().strip())
                        if item_info:
                            if not row_vals.get(12): row_vals[12] = item_info.get("code", "")
                            if not row_vals.get(13): row_vals[13] = item_info.get("unit", "")

                self._pnmh_rows.append(row_vals)
                self._pnmh_row_indices.append(row_idx)
                self._pnmh_image_filenames.append(str(row_vals.get(28, "") or ""))
            wb.close()
            self._populate_pnmh_table()
            self._refresh_original_row_snapshots()
        except Exception as e:
            QMessageBox.warning(self, "Lỗi đọc PNMH", str(e))

    def _populate_pnmh_table(self):
        self.pnmh_table.setRowCount(len(self._pnmh_rows))
        
        current_voucher = None
        current_bg = QColor("#0D2137")
        alt_bg = QColor("#0A1929")

        for row_idx, row_data in enumerate(self._pnmh_rows):
            voucher = row_data.get(2, "")
            if voucher != current_voucher:
                current_voucher = voucher
                current_bg, alt_bg = alt_bg, current_bg

            chk = QCheckBox()
            # Tăng kích thước vùng tick
            chk.setStyleSheet("QCheckBox::indicator { width: 22px; height: 22px; }")
            chk.stateChanged.connect(self._on_pnmh_check_changed)
            
            chk_widget = ClickableCheckBoxWidget(chk, bg_color=current_bg)
            self.pnmh_table.setCellWidget(row_idx, 0, chk_widget)

            for ui_col_idx, (pnmh_col, _) in enumerate(_PNMH_COLS_INFO, start=1):
                val = row_data.get(pnmh_col, "")
                if val is None: val = ""
                
                display_text = str(val)
                if pnmh_col in (16, 17, 20):
                    display_text = _fmt_money(val)
                elif pnmh_col in (1, 3): # NGÀY GHI SỔ (1) hoặc NGÀY HOÁ ĐƠN (3)
                    display_text = _fmt_date(val)

                item = QTableWidgetItem(display_text)
                item.setForeground(_COLOR_TEXT)
                item.setBackground(current_bg)
                
                # Align Right cho cột số
                if pnmh_col in (16, 17, 20):
                    item.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
                else:
                    item.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
                    
                self.pnmh_table.setItem(row_idx, ui_col_idx, item)

            self.pnmh_table.setRowHeight(row_idx, 30)
            
        self.pnmh_table.resizeColumnsToContents()
        # Ép độ rộng để tiết kiệm diện tích nhưng vẫn đủ nhìn
        self.pnmh_table.setColumnWidth(0, 50)   # X (Mở rộng một chút để dễ click)
        self.pnmh_table.setColumnWidth(1, 100)  # Số chứng từ
        self.pnmh_table.setColumnWidth(2, 90)   # Ngày hoá đơn
        self.pnmh_table.setColumnWidth(3, 70)   # Bộ phận
        self.pnmh_table.setColumnWidth(4, 120)  # Đối tượng
        self.pnmh_table.setColumnWidth(5, 300)  # Diễn giải (Giảm còn 1/2)
        self.pnmh_table.setColumnWidth(6, 100)  # Kho nhập
        self.pnmh_table.setColumnWidth(7, 100)  # Mã vật tư
        self.pnmh_table.setColumnWidth(8, 90)   # ĐVT
        self.pnmh_table.setColumnWidth(9, 80)   # Số lượng
        self.pnmh_table.setColumnWidth(10, 100) # Đơn giá
        self.pnmh_table.setColumnWidth(11, 120) # Thành tiền
        self.pnmh_table.setColumnWidth(12, 300) # Ghi chú (Note)

        # Luôn có 1 dòng trống dưới cùng
        self.pnmh_table.blockSignals(True)
        self._do_add_pnmh_row()
        self.pnmh_table.blockSignals(False)

    def _load_chiphi(self):
        self.chiphi_table.setRowCount(0)
        self._chiphi_rows.clear()
        self._chiphi_row_indices.clear()
        self._chiphi_image_filenames.clear()

        if not os.path.exists(self.chiphi_path): return

        try:
            wb = openpyxl.load_workbook(self.chiphi_path, data_only=True)
            ws = wb.active
            for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
                if not any(row): continue
                row_vals = {c_idx+1: val for c_idx, val in enumerate(row)}
                self._chiphi_rows.append(row_vals)
                self._chiphi_row_indices.append(row_idx)
                self._chiphi_image_filenames.append(str(row_vals.get(16, "") or ""))
            wb.close()
            self._populate_chiphi_table()
            self._refresh_original_row_snapshots()
        except Exception as e:
            QMessageBox.warning(self, "Lỗi đọc Chi phí", str(e))

    def _refresh_original_row_snapshots(self):
        self._pnmh_original_by_ws_row = {
            ws_row: copy.deepcopy(row_data)
            for ws_row, row_data in zip(self._pnmh_row_indices, self._pnmh_rows)
        }
        self._chiphi_original_by_ws_row = {
            ws_row: copy.deepcopy(row_data)
            for ws_row, row_data in zip(self._chiphi_row_indices, self._chiphi_rows)
        }

    def _populate_chiphi_table(self):
        self.chiphi_table.setRowCount(len(self._chiphi_rows))
        for row_idx, row_data in enumerate(self._chiphi_rows):
            for ui_col_idx, (cp_col, _) in enumerate(_CHIPHI_COLS_INFO):
                val = row_data.get(cp_col, "")
                if val is None: val = ""
                
                display_text = str(val)
                if cp_col == 14:
                    display_text = _fmt_money(val)
                elif cp_col in (1, 2): # SỐ CHỨNG TỪ (1) hoặc NGÀY GHI SỔ (2)
                    display_text = _fmt_date(val)

                item = QTableWidgetItem(display_text)
                item.setForeground(_COLOR_TEXT)
                
                if cp_col == 14:
                    item.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
                else:
                    item.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
                    
                self.chiphi_table.setItem(row_idx, ui_col_idx, item)
            self.chiphi_table.setRowHeight(row_idx, 30)
        self.chiphi_table.resizeColumnsToContents()
        # Ép độ rộng cho các cột trong Chi phí
        self.chiphi_table.setColumnWidth(0, 90)   # Ngày ghi sổ
        self.chiphi_table.setColumnWidth(1, 100)  # Số chứng từ
        self.chiphi_table.setColumnWidth(2, 70)   # Bộ phận
        self.chiphi_table.setColumnWidth(3, 120)  # Đối tượng
        self.chiphi_table.setColumnWidth(4, 300)  # Diễn giải (Giảm còn 1/2)
        self.chiphi_table.setColumnWidth(5, 70)   # TK Nợ
        self.chiphi_table.setColumnWidth(6, 70)   # TK Có
        self.chiphi_table.setColumnWidth(7, 120)  # Thành tiền
        self.chiphi_table.setColumnWidth(8, 300)  # Ghi chú (Note)

        # Luôn có 1 dòng trống dưới cùng (không có ảnh)
        self._chiphi_image_filenames.append("")
        self.chiphi_table.blockSignals(True)
        self._do_add_chiphi_row()
        self.chiphi_table.blockSignals(False)

    # ── Logic ──
    def _on_chiphi_cell_changed(self, row, col):
        self.chiphi_table.blockSignals(True)
        try:
            if col == 0: # Cột Ngày ghi sổ
                item = self.chiphi_table.item(row, col)
                if item:
                    item.setText(_fmt_date(item.text()))
                    
            # Tự động chuyển Tên NCC -> Mã NCC
            if col == 3: # Cột Mã đối tượng
                item = self.chiphi_table.item(row, col)
                if item:
                    text = item.text().strip()
                    if text and self.app_data:
                        for code, name in self.app_data.suppliers_dict.items():
                            if text.lower() == str(name).lower():
                                item.setText(str(code))
                                break
            
            # Tự động điền TK Nợ (6421) và TK Có (331) khi điền xong Diễn giải
            if col == 4: # Cột Diễn giải
                item = self.chiphi_table.item(row, col)
                if item and item.text().strip():
                    it_no = self.chiphi_table.item(row, 5)
                    if it_no and not it_no.text().strip():
                        it_no.setText("6421")
                    it_co = self.chiphi_table.item(row, 6)
                    if it_co and not it_co.text().strip():
                        it_co.setText("331")
        finally:
            self.chiphi_table.blockSignals(False)

        if row == self.chiphi_table.rowCount() - 1:
            has_text = False
            for c in range(self.chiphi_table.columnCount()):
                it = self.chiphi_table.item(row, c)
                if it and it.text().strip():
                    has_text = True
                    break
            if has_text:
                self.chiphi_table.blockSignals(True)
                self._do_add_chiphi_row()
                self.chiphi_table.blockSignals(False)

    def _on_pnmh_cell_changed(self, row, col):
        # col 5: Diễn giải, 7: Mã VT, 8: ĐVT, 9: SL, 10: Đơn giá, 11: Thành tiền
        self.pnmh_table.blockSignals(True)
        try:
            # 0. Format Ngày
            if col == 2: # Cột Ngày hoá đơn
                item = self.pnmh_table.item(row, col)
                if item:
                    item.setText(_fmt_date(item.text()))

            # 0.5 Tự động chuyển Tên NCC -> Mã NCC
            if col == 4: # Cột Đối tượng
                item = self.pnmh_table.item(row, col)
                if item:
                    text = item.text().strip()
                    if text and self.app_data:
                        for code, name in self.app_data.suppliers_dict.items():
                            if text.lower() == str(name).lower():
                                item.setText(str(code))
                                break

            # 1. Logic Lookup (Diễn giải -> Mã, ĐVT)
            if col == 5:
                text = self.pnmh_table.item(row, 5).text().lower().strip()
                if self.app_data and text in self.app_data.items_dict:
                    info = self.app_data.items_dict[text]
                    self.pnmh_table.item(row, 7).setText(info.get("code", ""))
                    self.pnmh_table.item(row, 8).setText(info.get("unit", ""))

            # 2. Logic Tính toán (SL, Đơn giá, Thành tiền)
            def _get_val(c):
                t = self.pnmh_table.item(row, c).text().replace(',', '')
                try: return float(t)
                except: return 0.0

            if col in (9, 10): # SL hoặc Đơn giá thay đổi -> Tính Thành tiền
                sl = _get_val(9)
                dg = _get_val(10)
                tt = sl * dg
                self.pnmh_table.item(row, 11).setText(_fmt_money(tt))
            elif col == 11: # Thành tiền thay đổi -> Tính ngược Đơn giá (nếu có SL)
                tt = _get_val(11)
                sl = _get_val(9)
                if sl != 0:
                    dg = tt / sl
                    self.pnmh_table.item(row, 10).setText(_fmt_money(dg))

        finally:
            if row == self.pnmh_table.rowCount() - 1:
                has_text = False
                for c in range(1, self.pnmh_table.columnCount()):
                    it = self.pnmh_table.item(row, c)
                    if it and it.text().strip():
                        has_text = True
                        break
                if has_text:
                    self._do_add_pnmh_row()
            self.pnmh_table.blockSignals(False)

    def _show_pnmh_context_menu(self, pos):
        from PyQt5.QtWidgets import QMenu
        index = self.pnmh_table.indexAt(pos)
        if not index.isValid():
            return

        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu { background-color: #0A2740; color: #BDD8E9; border: 1px solid #49769F; }
            QMenu::item:selected { background-color: #0A4174; color: #ffffff; }
        """)

        row = index.row()
        act_alias = menu.addAction("📝  Lưu vào Từ điển Alias")
        act_alias.triggered.connect(lambda: self._do_update_alias(row))
        menu.addSeparator()
        act_hard_case = menu.addAction("🚨  Báo cáo lỗi / Hard Case")
        act_hard_case.triggered.connect(lambda: self._collect_hard_case("PNMH", row))

        menu.exec_(self.pnmh_table.viewport().mapToGlobal(pos))

    def _show_chiphi_context_menu(self, pos):
        from PyQt5.QtWidgets import QMenu
        index = self.chiphi_table.indexAt(pos)
        if not index.isValid():
            return

        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu { background-color: #0A2740; color: #BDD8E9; border: 1px solid #49769F; }
            QMenu::item:selected { background-color: #0A4174; color: #ffffff; }
        """)

        row = index.row()
        act_hard_case = menu.addAction("🚨  Báo cáo lỗi / Hard Case")
        act_hard_case.triggered.connect(lambda: self._collect_hard_case("ChiPhi", row))

        menu.exec_(self.chiphi_table.viewport().mapToGlobal(pos))

    def _do_update_alias(self, row):
        it_ocr_name = self.pnmh_table.item(row, 5)
        it_item_code = self.pnmh_table.item(row, 7)
        it_alias_unit = self.pnmh_table.item(row, 8)
        it_price = self.pnmh_table.item(row, 10)
        it_note = self.pnmh_table.item(row, 12)

        if not it_ocr_name or not it_item_code:
            QMessageBox.warning(self, "Thiếu thông tin", "Dòng này chưa có Tên hàng hoặc Mã vật tư.")
            return

        ocr_name = it_ocr_name.text().strip()    # what the invoice/OCR says → saved as Alias
        item_code = it_item_code.text().strip()   # item code from master list
        alias_unit = it_alias_unit.text().strip() if it_alias_unit else ""

        if not ocr_name or not item_code:
            QMessageBox.warning(self, "Thiếu thông tin", "Tên hàng hoặc Mã vật tư đang để trống.")
            return

        # Look up the standard item name and master unit from the master list
        from data_manager import app_data
        ref_name = ""
        master_unit = ""
        rec = getattr(app_data, 'items_by_code', {}).get(item_code.lower(), {})
        if rec:
            ref_name = rec.get('name', '')
            master_unit = rec.get('unit', '')

        # Tính hệ số quy đổi dựa trên đơn giá
        note = it_note.text() if it_note else ""
        alias_factor = 1.0
        match = re.search(r"Gia unit:\s*([\d\s]+)", note)
        if match:
            try:
                base_price = float(match.group(1).replace(" ", ""))
                final_price = float(it_price.text().replace(",", "")) if it_price else 0
                if base_price > 0 and final_price > 0:
                    alias_factor = round(final_price / base_price, 3)
            except:
                alias_factor = 1.0

        confirm = QMessageBox.question(
            self, "Xác nhận",
            f"Lưu Alias mới?\n\n"
            f"• Alias (tên trên hóa đơn / OCR): {ocr_name}\n"
            f"• Tên chuẩn (trong hệ thống):      {ref_name or '(không tìm thấy)'}\n"
            f"• Mã VT: {item_code}\n"
            f"• ĐVT kho: {master_unit}\n"
            f"• ĐVT lóng: {alias_unit}\n"
            f"• Hệ số: {alias_factor}",
            QMessageBox.Yes | QMessageBox.No
        )

        if confirm == QMessageBox.Yes:
            try:
                app_data.save_alias(
                    alias=ocr_name,
                    code=item_code,
                    ref_name=ref_name,
                    master_unit=master_unit,
                    alias_unit=alias_unit,
                    alias_factor=alias_factor,
                )
                QMessageBox.information(
                    self, "Thành công",
                    f"Đã lưu:\n  '{ocr_name}'  →  '{ref_name or item_code}'"
                )
            except Exception as e:
                QMessageBox.critical(self, "Lỗi", f"Không thể lưu alias: {e}")

    def _collect_hard_case(self, tab_name: str, row: int):
        try:
            if tab_name == "PNMH":
                if row < 0 or row >= len(self._pnmh_row_indices):
                    QMessageBox.critical(self, "Lỗi", "Dữ liệu dòng bị lệch. Vui lòng tải lại dữ liệu.")
                    return
                ws_row = self._pnmh_row_indices[row]
                before_data = self._pnmh_original_by_ws_row.get(ws_row)
                if before_data is None:
                    QMessageBox.critical(self, "Lỗi", "Dữ liệu dòng bị lệch. Vui lòng tải lại dữ liệu.")
                    return
                after_data = self._read_pnmh_row_from_table(row)
                image_name = self._pnmh_image_filenames[row] if row < len(self._pnmh_image_filenames) else ""
            else:
                if row < 0 or row >= len(self._chiphi_row_indices):
                    QMessageBox.critical(self, "Lỗi", "Dữ liệu dòng bị lệch. Vui lòng tải lại dữ liệu.")
                    return
                ws_row = self._chiphi_row_indices[row]
                before_data = self._chiphi_original_by_ws_row.get(ws_row)
                if before_data is None:
                    QMessageBox.critical(self, "Lỗi", "Dữ liệu dòng bị lệch. Vui lòng tải lại dữ liệu.")
                    return
                after_data = self._read_chiphi_row_from_table(row)
                image_name = self._chiphi_image_filenames[row] if row < len(self._chiphi_image_filenames) else ""

            from PyQt5.QtWidgets import QInputDialog
            note, ok = QInputDialog.getText(
                self, "Báo cáo Hard Case",
                "Ghi chú về trường hợp này (không bắt buộc):\n"
                "VD: 'tên hàng không map được', 'đơn giá sai', 'mã vật tư nhầm'..."
            )
            if not ok:
                return

            def _json_safe(row_data: dict) -> dict:
                out = {}
                for key, value in row_data.items():
                    out[str(key)] = value.isoformat() if hasattr(value, "isoformat") else value
                return out

            hard_case_root = os.path.join(get_root_dir(), "HARD CASE COLLECTED")
            now      = datetime.datetime.now()
            date_dir = os.path.join(hard_case_root, now.strftime("%Y%m%d"))
            os.makedirs(date_dir, exist_ok=True)

            ts        = now.strftime("%Y%m%d_%H%M%S")
            base_name = f"{ts}_{tab_name.lower()}_r{row + 1}"

            copied_images: list[str] = []
            if image_name:
                src_path = os.path.join(os.path.dirname(self.pnmh_path), image_name)
                if os.path.exists(src_path):
                    ext      = os.path.splitext(image_name)[1].lower()
                    img_name = f"{base_name}{ext}"
                    shutil.copy2(src_path, os.path.join(date_dir, img_name))
                    copied_images.append(img_name)

            report = {
                "created_at": now.isoformat(timespec="seconds"),
                "tab": tab_name,
                "table_row_index_0_based": row,
                "table_row_index_1_based": row + 1,
                "worksheet_row_index": ws_row,
                "pnmh_file": os.path.basename(self.pnmh_path or ""),
                "chiphi_file": os.path.basename(self.chiphi_path or ""),
                "image_paths": copied_images,
                "note": note.strip(),
                "before": _json_safe(before_data),
                "after": _json_safe(after_data),
            }

            with open(os.path.join(date_dir, f"{base_name}.json"), "w", encoding="utf-8") as f:
                json.dump(report, f, ensure_ascii=False, indent=2)

            QMessageBox.information(
                self,
                "Thành công",
                f"Đã lưu Hard Case:\n{base_name}"
            )
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể lưu Hard Case:\n{e}")

    def _on_pnmh_check_changed(self):
        selected = self._get_pnmh_selected_indices()
        self.btn_transfer.setEnabled(len(selected) > 0)

    def _get_pnmh_selected_indices(self) -> list[int]:
        selected = []
        for r in range(self.pnmh_table.rowCount()):
            chk_widget = self.pnmh_table.cellWidget(r, 0)
            if chk_widget:
                chk = chk_widget.findChild(QCheckBox)
                if chk and chk.isChecked():
                    selected.append(r)
        return selected

    def _read_pnmh_row_from_table(self, row_idx: int) -> dict:
        data = dict(self._pnmh_rows[row_idx]) if row_idx < len(self._pnmh_rows) else {}
        for ui_col_idx, (pnmh_col, _) in enumerate(_PNMH_COLS_INFO, start=1):
            item = self.pnmh_table.item(row_idx, ui_col_idx)
            if item:
                raw = item.text().strip()
                if pnmh_col in (16, 17, 20):
                    try: data[pnmh_col] = float(raw.replace(",", "")) if raw else None
                    except: data[pnmh_col] = None
                else:
                    data[pnmh_col] = raw or None
        return data
        
    def _read_chiphi_row_from_table(self, row_idx: int) -> dict:
        data = dict(self._chiphi_rows[row_idx]) if row_idx < len(self._chiphi_rows) else {}
        for ui_col_idx, (cp_col, _) in enumerate(_CHIPHI_COLS_INFO):
            item = self.chiphi_table.item(row_idx, ui_col_idx)
            if item:
                raw = item.text().strip()
                if cp_col == 14:
                    try: data[cp_col] = float(raw.replace(",", "")) if raw else None
                    except: data[cp_col] = None
                else:
                    data[cp_col] = raw or None
        return data

    def _do_add_pnmh_row(self):
        row_idx = self.pnmh_table.rowCount()
        self.pnmh_table.insertRow(row_idx)
        self.pnmh_table.setRowHeight(row_idx, 30)
        
        # Checkbox (Sử dụng ClickableCheckBoxWidget để tăng diện tích click)
        chk = QCheckBox()
        chk.setStyleSheet("QCheckBox::indicator { width: 22px; height: 22px; }")
        chk_widget = ClickableCheckBoxWidget(chk)
        self.pnmh_table.setCellWidget(row_idx, 0, chk_widget)
        chk.stateChanged.connect(self._on_pnmh_check_changed)
        
        # Items mặc định trống
        for c in range(1, self.pnmh_table.columnCount()):
            item = QTableWidgetItem("")
            item.setForeground(_COLOR_TEXT)
            self.pnmh_table.setItem(row_idx, c, item)
            
        # Không thêm vào self._pnmh_rows và self._pnmh_row_indices ở đây để báo hiệu là dòng mới

    def _do_add_chiphi_row(self):
        row_idx = self.chiphi_table.rowCount()
        self.chiphi_table.insertRow(row_idx)
        self.chiphi_table.setRowHeight(row_idx, 30)
        for c in range(self.chiphi_table.columnCount()):
            item = QTableWidgetItem("")
            item.setForeground(_COLOR_TEXT)
            self.chiphi_table.setItem(row_idx, c, item)

    def _do_delete(self):
        active_tab = self.tabs.currentIndex()
        if active_tab == 0: # PNMH
            selected = self.pnmh_table.selectedRanges()
            if not selected:
                QMessageBox.information(self, "Thông báo", "Vui lòng quét chọn các ô trong dòng cần xoá.")
                return
            
            rows_to_del = set()
            for rng in selected:
                for r in range(rng.topRow(), rng.bottomRow() + 1):
                    rows_to_del.add(r)
            
            if not rows_to_del: return
            
            if QMessageBox.question(self, "Xác nhận", f"Xoá {len(rows_to_del)} dòng đang chọn?") == QMessageBox.Yes:
                for r in sorted(list(rows_to_del), reverse=True):
                    self.pnmh_table.removeRow(r)
                    if r < len(self._pnmh_rows):
                        self._deleted_pnmh_ws_rows.append(self._pnmh_row_indices[r])
                        self._pnmh_rows.pop(r)
                        self._pnmh_row_indices.pop(r)
                        if r < len(self._pnmh_image_filenames):  # shorter: manually-added rows have no entry here
                            self._pnmh_image_filenames.pop(r)
                assert len(self._pnmh_image_filenames) == len(self._pnmh_rows), \
                    "Image filename list desynced from rows list"
        else: # Chi phí
            selected = self.chiphi_table.selectedRanges()
            if not selected:
                QMessageBox.information(self, "Thông báo", "Vui lòng quét chọn các ô trong dòng cần xoá.")
                return
            
            rows_to_del = set()
            for rng in selected:
                for r in range(rng.topRow(), rng.bottomRow() + 1):
                    rows_to_del.add(r)
            
            if not rows_to_del: return
            
            if QMessageBox.question(self, "Xác nhận", f"Xoá {len(rows_to_del)} dòng đang chọn?") == QMessageBox.Yes:
                for r in sorted(list(rows_to_del), reverse=True):
                    self.chiphi_table.removeRow(r)
                    if r < len(self._chiphi_rows):
                        self._deleted_chiphi_ws_rows.append(self._chiphi_row_indices[r])
                        self._chiphi_rows.pop(r)
                        self._chiphi_row_indices.pop(r)

    def _do_transfer(self):
        selected_ui_indices = self._get_pnmh_selected_indices()
        if not selected_ui_indices: return

        if not os.path.exists(self.chiphi_path):
            if not os.path.exists(_CHIPHI_TEMPLATE_FILE):
                QMessageBox.critical(self, "Lỗi", f"Không tìm thấy template:\n{_CHIPHI_TEMPLATE_FILE}")
                return
            shutil.copy(_CHIPHI_TEMPLATE_FILE, self.chiphi_path)

        try:
            chiphi_wb = openpyxl.load_workbook(self.chiphi_path)
            chiphi_ws = chiphi_wb.active
            chiphi_row = _find_first_empty_row(chiphi_ws)

            for ui_idx in selected_ui_indices:
                row_data = self._read_pnmh_row_from_table(ui_idx)
                src_fn = self._pnmh_image_filenames[ui_idx] if ui_idx < len(self._pnmh_image_filenames) else ""
                inv_date = str(row_data.get(3, "") or "")
                _append_to_chiphi(chiphi_ws, chiphi_row, row_data, self.kho_dict, source_filename=src_fn, invoice_date=inv_date)
                chiphi_row += 1

            chiphi_wb.save(self.chiphi_path)
            chiphi_wb.close()

            pnmh_wb = openpyxl.load_workbook(self.pnmh_path)
            pnmh_ws = pnmh_wb.active

            ws_rows_to_delete = sorted([self._pnmh_row_indices[i] for i in selected_ui_indices], reverse=True)
            for ws_row in ws_rows_to_delete:
                pnmh_ws.delete_rows(ws_row)
                # Cập nhật lại indices trong RAM để không bị lệch do xoá dòng
                for i in range(len(self._pnmh_row_indices)):
                    if self._pnmh_row_indices[i] > ws_row:
                        self._pnmh_row_indices[i] -= 1
                for i in range(len(self._deleted_pnmh_ws_rows)):
                    if self._deleted_pnmh_ws_rows[i] > ws_row:
                        self._deleted_pnmh_ws_rows[i] -= 1

            pnmh_wb.save(self.pnmh_path)
            pnmh_wb.close()

            for ui_idx in sorted(selected_ui_indices, reverse=True):
                self.pnmh_table.removeRow(ui_idx)
                self._pnmh_rows.pop(ui_idx)
                self._pnmh_row_indices.pop(ui_idx)
                if ui_idx < len(self._pnmh_image_filenames):  # shorter: manually-added rows have no entry here
                    self._pnmh_image_filenames.pop(ui_idx)
            assert len(self._pnmh_image_filenames) == len(self._pnmh_rows), \
                "Image filename list desynced from rows list"
            self._on_pnmh_check_changed()
            
            self._load_chiphi()
            self.tabs.setCurrentWidget(self.tab_chiphi)

        except PermissionError as e:
            QMessageBox.critical(self, "Lỗi", f"File đang mở trong Excel:\n{e}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def _do_save_all_and_close(self):
        try:
            if os.path.exists(self.pnmh_path):
                wb = openpyxl.load_workbook(self.pnmh_path)
                ws = wb.active
                # Xóa dữ liệu cũ trong vùng làm việc (giả sử từ dòng 3)
                # Nhưng thực tế chúng ta delete_rows khi transfer, nên ở đây chỉ cần ghi đè và append
                
                # 1. Ghi đè các dòng hiện có
                for ui_idx, ws_row in enumerate(self._pnmh_row_indices):
                    row_data = self._read_pnmh_row_from_table(ui_idx)
                    for col, val in row_data.items():
                        ws.cell(row=ws_row, column=col, value=val)
                
                # 2. Append các dòng mới thêm
                if self.pnmh_table.rowCount() > len(self._pnmh_row_indices):
                    next_row = _find_first_empty_row(ws)
                    for ui_idx in range(len(self._pnmh_row_indices), self.pnmh_table.rowCount()):
                        row_data = self._read_pnmh_row_from_table(ui_idx)
                        if not any(v for v in row_data.values() if v is not None and str(v).strip()):
                            continue
                        for col, val in row_data.items():
                            ws.cell(row=next_row, column=col, value=val)
                        next_row += 1

                # 3. Thực hiện xoá các dòng đã bị mark delete (xoá từ dưới lên trên)
                for ws_row in sorted(self._deleted_pnmh_ws_rows, reverse=True):
                    ws.delete_rows(ws_row)

                wb.save(self.pnmh_path)
                wb.close()

            if os.path.exists(self.chiphi_path):
                wb = openpyxl.load_workbook(self.chiphi_path)
                ws = wb.active
                
                # 1. Ghi đè các dòng hiện có
                for ui_idx, ws_row in enumerate(self._chiphi_row_indices):
                    row_data = self._read_chiphi_row_from_table(ui_idx)
                    for col, val in row_data.items():
                        ws.cell(row=ws_row, column=col, value=val)
                
                # 2. Append các dòng mới thêm
                if self.chiphi_table.rowCount() > len(self._chiphi_row_indices):
                    next_row = _find_first_empty_row(ws)
                    for ui_idx in range(len(self._chiphi_row_indices), self.chiphi_table.rowCount()):
                        row_data = self._read_chiphi_row_from_table(ui_idx)
                        if not any(v for v in row_data.values() if v is not None and str(v).strip()):
                            continue
                        for col, val in row_data.items():
                            ws.cell(row=next_row, column=col, value=val)
                        next_row += 1

                # 3. Thực hiện xoá các dòng đã bị mark delete (xoá từ dưới lên trên)
                for ws_row in sorted(self._deleted_chiphi_ws_rows, reverse=True):
                    ws.delete_rows(ws_row)

                wb.save(self.chiphi_path)
                wb.close()

            self.accept()
        except PermissionError:
            QMessageBox.critical(self, "Lỗi", "File đang được mở bởi Excel. Vui lòng đóng Excel và thử lại.")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def _stylesheet(self) -> str:
        return """
        QDialog {
            background: qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #001D39, stop:1 #000E1F);
        }
        QLabel { color: #BDD8E9; }
        QTabWidget::pane { border: 1px solid rgba(123,189,232,0.3); border-radius: 4px; background: transparent; }
        QTabBar::tab {
            background: rgba(0,29,57,0.6); color: #6EA2B3;
            border: 1px solid rgba(123,189,232,0.2); border-bottom-color: transparent;
            padding: 8px 16px; margin-right: 2px;
            border-top-left-radius: 4px; border-top-right-radius: 4px;
        }
        QTabBar::tab:selected { background: #0A4174; color: #ffffff; border-color: rgba(123,189,232,0.5); }
        QTableWidget {
            background-color: transparent;
            border: none;
            gridline-color: rgba(123,189,232,0.1);
            color: #BDD8E9;
            font-size: 13px;
        }
        QTableWidget::item:selected { background-color: rgba(10,65,116,0.6); color: #ffffff; }
        QHeaderView::section {
            background-color: #0A2740; color: #7BBDE8; font-weight: 700; font-size: 12px;
            border: 1px solid rgba(123,189,232,0.15); padding: 6px;
        }
        QPushButton { border-radius: 8px; padding: 8px 18px; font-size: 13px; font-weight: 600; min-height: 32px; }
        QPushButton#btn_transfer {
            background: qlineargradient(x1:0,y1:0,x2:1,y2:1, stop:0 #0A4174, stop:1 #49769F);
            color: #ffffff; border: 1px solid rgba(123,189,232,0.4);
        }
        QPushButton#btn_transfer:hover { background: qlineargradient(x1:0,y1:0,x2:1,y2:1, stop:0 #49769F, stop:1 #7BBDE8); }
        QPushButton#btn_transfer:disabled { background: rgba(10,65,116,0.25); color: #49769F; }
        QPushButton#btn_save {
            background-color: rgba(0,60,0,0.4); color: #86efac; border: 1px solid rgba(100,200,100,0.35);
        }
        QPushButton#btn_save:hover { background-color: rgba(0,120,0,0.4); }
        QPushButton#btn_delete {
            background-color: rgba(100,0,0,0.4); color: #fca5a5; border: 1px solid rgba(200,100,100,0.3);
        }
        QPushButton#btn_delete:hover { background-color: rgba(160,0,0,0.5); }
        QScrollBar:vertical { background: rgba(0,13,26,0.5); width: 8px; border-radius: 4px; }
        QScrollBar::handle:vertical { background: #49769F; border-radius: 4px; min-height: 30px; }
        QScrollBar:horizontal { background: rgba(0,13,26,0.5); height: 8px; border-radius: 4px; }
        QScrollBar::handle:horizontal { background: #49769F; border-radius: 4px; min-width: 30px; }
        """
