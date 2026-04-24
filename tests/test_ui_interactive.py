"""
Interactive GUI Test for Lighthouse OCR v6.0
This script tests UI components like the PostProcessDialog without requiring a full OCR run.
"""
import sys
import os
from PyQt5.QtWidgets import QApplication, QDialog
from PyQt5.QtCore import Qt

# Add root to sys.path
sys.path.insert(0, os.path.abspath('.'))

from post_process_dialog import PostProcessDialog
from data_manager import DataManager

def create_mock_excel(pnmh_path, chiphi_path):
    import openpyxl
    # Create PNMH mock
    wb = openpyxl.Workbook()
    ws = wb.active
    # Header row (min_row=3 logic in app)
    headers = ["STT", "Số CT", "Ngày", "Mã ĐT", "Đối tượng", "Diễn giải", "Mã Kho", "Mã VT", "ĐVT", "Số lượng", "Đơn giá", "Thành tiền", "Ghi chú"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    
    # Data row 1: Normal item
    ws.cell(row=4, column=2, value="V001")
    ws.cell(row=4, column=6, value="Dưa leo")
    ws.cell(row=4, column=10, value=2)
    ws.cell(row=4, column=11, value=15000)
    ws.cell(row=4, column=12, value=30000)
    
    # Data row 2: Missing lookup info
    ws.cell(row=5, column=2, value="V001")
    ws.cell(row=5, column=6, value="Cà chua")
    ws.cell(row=5, column=10, value=5)
    ws.cell(row=5, column=11, value=20000)
    ws.cell(row=5, column=12, value=100000)
    
    wb.save(pnmh_path)
    wb.close()
    
    # Create ChiPhi mock
    wb_cp = openpyxl.Workbook()
    ws_cp = wb_cp.active
    cp_headers = ["STT", "Ngày", "Số CT", "Bộ phận", "Đối tượng", "Diễn giải", "TK Nợ", "TK Có", "Thành tiền", "Ghi chú"]
    for i, h in enumerate(cp_headers, 1):
        ws_cp.cell(row=3, column=i, value=h)
    wb_cp.save(chiphi_path)
    wb_cp.close()

class MockAppData:
    def __init__(self):
        self.suppliers_dict = {"NCC01": "Lighthouse Supplier"}
        self.items_dict = {
            "dưa leo": {"code": "VT01", "name": "Dưa leo", "unit": "Kg", "dvt0": "Thùng"},
            "cà chua": {"code": "VT02", "name": "Cà chua", "unit": "Kg", "dvt0": "Rổ"}
        }
        self.kho_dict = {"BEP": "KHO-BEP", "BAR": "KHO-BAR"}

def main():
    app = QApplication(sys.argv)
    
    # Create temp folder for test files
    temp_dir = "TEMP_TEST"
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)
    
    pnmh_path = os.path.join(temp_dir, "TEST_PNMH.xlsx")
    chiphi_path = os.path.join(temp_dir, "TEST_ChiPhi.xlsx")
    
    print("Creating mock excel files...")
    create_mock_excel(pnmh_path, chiphi_path)
    
    print("Initializing mock data...")
    mock_data = MockAppData()
    
    print("Launching PostProcessDialog...")
    print("TEST CASE GUIDE:")
    print("1. [Manual Row] Click '+ Thêm dòng' in PNMH tab. Check if row appears.")
    print("2. [Lookup] Type 'Dưa leo' in Diễn giải of the new row. Check if Mã VT and ĐVT fill automatically.")
    print("3. [Math] Enter Số lượng=10 and Đơn giá=5000. Check if Thành tiền=50,000.")
    print("4. [Reverse Math] Change Thành tiền to 100,000. Check if Đơn giá updates to 10,000.")
    print("5. [DVT Dropdown] Click ĐVT cell for 'Dưa leo'. Check if 'Kg' and 'Thùng' appear in list.")
    print("6. [Delete] Select a row and click 'Xoá Dòng Đang Chọn'.")
    print("7. [Save] Click 'Lưu & Đóng'. Verify files in TEMP_TEST folder.")
    
    dlg = PostProcessDialog(pnmh_path, chiphi_path, mock_data)
    res = dlg.exec_()
    
    if res == QDialog.Accepted:
        print("\n✅ TEST SUCCESS: Changes saved.")
    else:
        print("\n🚫 TEST CANCELLED: Dialog closed without saving.")
        # Optional: verify files were deleted if it was a real run, 
        # but here the dialog rejection logic is in main_app_qt, not dialog itself.
        # So we can manually verify the files here if needed.

    print("\nClean up TEMP_TEST manually after checking.")

if __name__ == "__main__":
    main()
