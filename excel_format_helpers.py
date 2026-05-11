"""
excel_format_helpers.py
Formatting constants and helper functions for Excel output (openpyxl).
"""

from datetime import date, datetime
from typing import Any

from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ---- Color palette ----
_FILL_NONE      = PatternFill(fill_type=None)
_FILL_PALE_BLUE = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
_FILL_RED       = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
_FILL_YELLOW    = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_FILL_HEADER    = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")

_FONT_HEADER    = Font(color="FFFFFF", bold=True, size=10)
_FONT_DATA      = Font(size=14)
_THIN_BORDER    = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

_NOTES_CELL_COMMENT = (
    "QUY TẮC MÀU SẮC & TÍNH TOÁN:\n"
    "----------------------------------\n"
    "🔴 ĐỎ: Hóa đơn rủi ro cao (Confidence thấp hoặc thiếu NCC).\n"
    "🟡 VÀNG: Vật tư chưa khớp Master List hoặc Sai số liệu (Qty * Price != Total).\n"
    "🔵 XANH NHẠT: Hóa đơn bình thường (tô xen kẽ để dễ nhìn).\n\n"
    "LOGIC TÍNH GIÁ:\n"
    "- Đơn giá (cột Q,R): Đã bao gồm VAT và Chiết khấu/Phí vận chuyển phân bổ.\n"
    "- Phân bổ: Chiết khấu/Phí cấp đơn được chia theo tỷ trọng giá trị từng dòng hàng.\n"
    "- VAT: Được tính trên giá sau chiết khấu.\n"
    "- Giá gốc: Xem chi tiết phép tính điều chỉnh trong nội dung ô ghi chú."
)


def _get_fill_for_invoice(invoice_idx: int, is_high_risk: bool) -> PatternFill:
    if is_high_risk:
        return _FILL_RED
    return _FILL_NONE if invoice_idx % 2 == 0 else _FILL_PALE_BLUE


def _apply_fill_to_row(ws, row_num: int, fill: PatternFill, max_col: int = 27):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.border = _THIN_BORDER
        cell.font = _FONT_DATA


def _apply_global_font(ws, size: int = 14):
    """Áp dụng cỡ chữ cho toàn bộ các ô có dữ liệu trong sheet, giới hạn vùng dùng được để tránh lag."""
    limit_row = min(ws.max_row, 1000)
    limit_col = min(ws.max_column, 50)

    for row in ws.iter_rows(min_row=1, max_row=limit_row, min_col=1, max_col=limit_col):
        for cell in row:
            if cell.font:
                cell.font = Font(
                    name=cell.font.name or "Calibri",
                    size=size,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    vertAlign=cell.font.vertAlign,
                    underline=cell.font.underline,
                    strike=cell.font.strike,
                    color=cell.font.color
                )
            else:
                cell.font = Font(size=size)

        row_num = row[0].row
        ws.row_dimensions[row_num].height = 30 if row_num <= 2 else 18


def _auto_fit_columns(ws, max_rows: int = 1000, max_cols: int = 50):
    """
    Tự động điều chỉnh độ rộng cột dựa trên nội dung dài nhất.
    Sử dụng hệ số nhân 1.2 cho font size 14 để tránh bị cắt chữ.
    """
    limit_row = min(ws.max_row, max_rows)
    limit_col = min(ws.max_column, max_cols)

    for col_idx in range(1, limit_col + 1):
        max_len = 0
        for row in ws.iter_rows(min_row=1, max_row=limit_row, min_col=col_idx, max_col=col_idx):
            cell = row[0]
            if cell.value is None:
                continue
            val = cell.value
            if isinstance(val, str) and val.startswith("="):
                continue
            text = str(val) if val is not None else ""
            current_len = len(text)
            if current_len > max_len:
                max_len = current_len

        if max_len <= 0:
            continue

        col_letter = get_column_letter(col_idx)
        target_width = min(70, max(10, (max_len + 2) * 1.15))
        ws.column_dimensions[col_letter].width = target_width


def _format_date(date_val: Any) -> str:
    """Chuẩn hoá ngày về dạng dd/mm/yyyy"""
    if not date_val:
        return ""
    if isinstance(date_val, (date, datetime)):
        return date_val.strftime("%d/%m/%Y")

    date_str = str(date_val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(date_str, fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return date_str


def _get_item_department(raw_group: str) -> str:
    group = str(raw_group or "").strip().upper()
    for dept_key in ("RANG", "BANH", "BAR", "BEP"):
        if group.startswith(dept_key):
            return dept_key
    return ""
